from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.models import User
from django.contrib import messages 
from django.contrib.auth.decorators import login_required, user_passes_test
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_POST
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.cache import never_cache
from django.db.models import Count, Q, Avg, Sum, F, DecimalField, ExpressionWrapper, Case, When, Value, IntegerField
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.contrib.admin.views.decorators import staff_member_required
from django.utils import timezone
from django.conf import settings
from django.urls import reverse
from django.core.mail import send_mail
from django.template.loader import render_to_string
from django.utils.http import urlsafe_base64_encode, urlsafe_base64_decode
from django.utils.encoding import force_bytes
from django.contrib.auth.tokens import default_token_generator
from django.conf import settings
from decimal import Decimal
from urllib.parse import urlencode
import logging

from .models import CategoryIcon, Slider, Feature, Banner, Product, DealCountdown, UserProfile, Cart, Wishlist, ProductImage, ProductReview, ReviewImage, ReviewVote, ProductQuestion, Order, OrderItem, OrderStatusHistory, AdminEmailSettings, ProductStockNotification, BrandPartner, SiteSettings, LoyaltyPoints, PointsTransaction
from .email_utils import send_order_confirmation_email, send_order_status_update_email, send_admin_order_notification

# ===== ADMIN PANEL VIEWS =====

@login_required(login_url='login')
@staff_member_required(login_url='login')
def admin_test(request):
    """Admin Test Page"""
    return render(request, 'admin_panel/test.html')

@login_required(login_url='login')
@staff_member_required(login_url='login')
def admin_widgets(request):
    """Admin Widgets Page"""
    return render(request, 'admin_panel/widgets.html')

@login_required(login_url='login')
@staff_member_required(login_url='login')
def admin_dashboard(request):
    """Admin Dashboard with comprehensive e-commerce statistics"""
    from datetime import timedelta
    from django.db.models.functions import TruncDate, TruncMonth, ExtractHour
    import calendar
    
    # Get current date for filtering
    today = timezone.now().date()
    last_7_days = today - timedelta(days=7)
    last_30_days = today - timedelta(days=30)
    current_month_start = today.replace(day=1)
    
    # Basic Statistics
    total_products = Product.objects.count()
    active_products = Product.objects.filter(is_active=True).count()
    total_users = User.objects.count()
    total_orders = Order.objects.count()

    # Paid orders (for consistent revenue calculations)
    paid_orders_qs = Order.objects.filter(payment_status='PAID')
    
    # Revenue Statistics
    total_revenue = paid_orders_qs.aggregate(total=Sum('total_amount'))['total'] or 0
    
    monthly_revenue = paid_orders_qs.filter(
        created_at__gte=current_month_start
    ).aggregate(total=Sum('total_amount'))['total'] or 0
    
    weekly_revenue = paid_orders_qs.filter(
        created_at__gte=last_7_days
    ).aggregate(total=Sum('total_amount'))['total'] or 0
    
    # Order Statistics
    pending_orders = Order.objects.filter(order_status='PENDING').count()
    processing_orders = Order.objects.filter(order_status='PROCESSING').count()
    shipped_orders = Order.objects.filter(order_status='SHIPPED').count()
    delivered_orders = Order.objects.filter(order_status='DELIVERED').count()
    
    # Recent Orders
    recent_orders = Order.objects.all().select_related('user').order_by('-created_at')[:10]

    # Recent Order Items for dashboard table (show first 7 recent orders with their items)
    recent_order_items = Order.objects.all().select_related('user').prefetch_related('items').order_by('-created_at')[:7]
    
    # Top Selling Products (by total quantity sold)
    top_products = (
        Product.objects
        .annotate(sales_count=Sum('orderitem__quantity'))
        .filter(sales_count__gt=0)
        .order_by('-sales_count')[:5]
    )
    
    # Low Stock Products
    low_stock_products = Product.objects.filter(stock__lte=10, is_active=True).order_by('stock')[:5]
    
    # Recent Reviews
    recent_reviews = ProductReview.objects.select_related('user', 'product').order_by('-created_at')[:5]
    
    # Daily Sales Chart Data (Last 7 Days)
    daily_sales = paid_orders_qs.filter(
        created_at__gte=last_7_days
    ).annotate(
        date=TruncDate('created_at')
    ).values('date').annotate(
        total=Sum('total_amount')
    ).order_by('date')

    # Visitors (new users) in last 7 days
    new_users_last_7 = User.objects.filter(date_joined__date__gte=last_7_days).count()
    prev_7_start = last_7_days - timedelta(days=7)
    new_users_prev_7 = User.objects.filter(
        date_joined__date__gte=prev_7_start,
        date_joined__date__lt=last_7_days
    ).count()
    visitors_weekly_percent = round((new_users_last_7 / new_users_prev_7) * 100, 1) if new_users_prev_7 else (100.0 if new_users_last_7 else 0.0)

    visitors_series_qs = User.objects.filter(date_joined__date__gte=last_7_days).annotate(
        date=TruncDate('date_joined')
    ).values('date').annotate(count=Count('id')).order_by('date')
    visitors_series_map = {row['date']: row['count'] for row in visitors_series_qs}
    visitors_week_labels = [today - timedelta(days=i) for i in range(6, -1, -1)]
    visitors_weekly_series = [visitors_series_map.get(day, 0) for day in visitors_week_labels]

    # Visitors (today, yesterday, last 30 days)
    yesterday = today - timedelta(days=1)

    visitors_today_qs = User.objects.filter(date_joined__date=today).annotate(
        hour=ExtractHour('date_joined')
    ).values('hour').annotate(count=Count('id')).order_by('hour')
    visitors_today_map = {row['hour']: row['count'] for row in visitors_today_qs}
    visitors_today_series = [visitors_today_map.get(h, 0) for h in range(24)]
    visitors_today_count = sum(visitors_today_series)

    visitors_yesterday_qs = User.objects.filter(date_joined__date=yesterday).annotate(
        hour=ExtractHour('date_joined')
    ).values('hour').annotate(count=Count('id')).order_by('hour')
    visitors_yesterday_map = {row['hour']: row['count'] for row in visitors_yesterday_qs}
    visitors_yesterday_series = [visitors_yesterday_map.get(h, 0) for h in range(24)]
    visitors_yesterday_count = sum(visitors_yesterday_series)

    last_30_start = today - timedelta(days=29)
    visitors_last_month_qs = User.objects.filter(date_joined__date__gte=last_30_start).annotate(
        date=TruncDate('date_joined')
    ).values('date').annotate(count=Count('id')).order_by('date')
    visitors_last_month_map = {row['date']: row['count'] for row in visitors_last_month_qs}
    visitors_last_month_labels = [last_30_start + timedelta(days=i) for i in range(30)]
    visitors_last_month_series = [visitors_last_month_map.get(day, 0) for day in visitors_last_month_labels]
    visitors_last_month_count = sum(visitors_last_month_series)

    # Activity (orders) last 7 days vs previous 7 days
    activity_last_7 = Order.objects.filter(created_at__date__gte=last_7_days).count()
    activity_prev_7 = Order.objects.filter(
        created_at__date__gte=prev_7_start,
        created_at__date__lt=last_7_days
    ).count()
    activity_weekly_percent = round((activity_last_7 / activity_prev_7) * 100, 1) if activity_prev_7 else (100.0 if activity_last_7 else 0.0)

    activity_series_qs = Order.objects.filter(created_at__date__gte=last_7_days).annotate(
        date=TruncDate('created_at')
    ).values('date').annotate(count=Count('id')).order_by('date')
    activity_series_map = {row['date']: row['count'] for row in activity_series_qs}
    activity_weekly_series = [activity_series_map.get(day, 0) for day in visitors_week_labels]

    # Sales (ranges)
    last_30_revenue = paid_orders_qs.filter(created_at__date__gte=last_30_days).aggregate(total=Sum('total_amount'))['total'] or 0
    sales_today = paid_orders_qs.filter(created_at__date=today).aggregate(total=Sum('total_amount'))['total'] or 0
    sales_yesterday = paid_orders_qs.filter(created_at__date=today - timedelta(days=1)).aggregate(total=Sum('total_amount'))['total'] or 0
    sales_last_month = last_30_revenue
    last_6_start = (current_month_start - timedelta(days=1)).replace(day=1) - timedelta(days=30 * 5)
    sales_last_6_months = paid_orders_qs.filter(created_at__date__gte=last_6_start).aggregate(total=Sum('total_amount'))['total'] or 0
    prev_30_start = last_30_days - timedelta(days=30)
    prev_30_revenue = paid_orders_qs.filter(
        created_at__date__gte=prev_30_start,
        created_at__date__lt=last_30_days
    ).aggregate(total=Sum('total_amount'))['total'] or 0
    sales_change_percent = round(((last_30_revenue - prev_30_revenue) / prev_30_revenue) * 100, 1) if prev_30_revenue else (100.0 if last_30_revenue else 0.0)

    # Profit & Expenses (estimated from revenue)
    profit_last_30 = last_30_revenue * Decimal('0.28')
    expenses_last_30 = last_30_revenue - profit_last_30
    expenses_percent = round((expenses_last_30 / last_30_revenue) * 100, 1) if last_30_revenue else 0
    expenses_remaining_percent = max(0, 100 - expenses_percent)

    # Transactions (paid orders) & change
    transactions_last_30 = paid_orders_qs.filter(created_at__date__gte=last_30_days).count()
    transactions_prev_30 = paid_orders_qs.filter(
        created_at__date__gte=prev_30_start,
        created_at__date__lt=last_30_days
    ).count()
    transactions_change_percent = round(((transactions_last_30 - transactions_prev_30) / transactions_prev_30) * 100, 1) if transactions_prev_30 else (100.0 if transactions_last_30 else 0.0)

    # Yearly income/expense overview
    current_year_start = today.replace(month=1, day=1)
    income_by_month_qs = paid_orders_qs.filter(created_at__date__gte=current_year_start).annotate(
        month=TruncMonth('created_at')
    ).values('month').annotate(total=Sum('total_amount')).order_by('month')
    income_month_map = {row['month'].month: row['total'] for row in income_by_month_qs if row['month']}
    month_labels = [calendar.month_abbr[i] for i in range(1, 13)]
    income_yearly_series = [float(income_month_map.get(i, 0) or 0) for i in range(1, 13)]
    expense_yearly_series = [float((income_month_map.get(i, 0) or 0) * Decimal('0.72')) for i in range(1, 13)]

    # Monthly report summary (current month vs previous)
    previous_month_start = (current_month_start - timedelta(days=1)).replace(day=1)
    previous_month_end = current_month_start - timedelta(days=1)
    prev_month_revenue = paid_orders_qs.filter(
        created_at__date__gte=previous_month_start,
        created_at__date__lte=previous_month_end
    ).aggregate(total=Sum('total_amount'))['total'] or 0
    report_income = monthly_revenue
    report_expense = monthly_revenue * Decimal('0.72')
    report_profit = monthly_revenue * Decimal('0.28')
    report_income_change = round(((monthly_revenue - prev_month_revenue) / prev_month_revenue) * 100, 1) if prev_month_revenue else (100.0 if monthly_revenue else 0.0)

    # Performance (last 6 months)
    last_6_months = [((current_month_start - timedelta(days=1)).replace(day=1) - timedelta(days=30 * i)) for i in range(5, -1, -1)]
    perf_month_labels = [calendar.month_abbr[m.month] for m in last_6_months]
    perf_income_series = []
    for m in last_6_months:
        next_m = (m + timedelta(days=32)).replace(day=1)
        total = paid_orders_qs.filter(created_at__date__gte=m, created_at__date__lt=next_m).aggregate(total=Sum('total_amount'))['total'] or 0
        perf_income_series.append(float(total))
    perf_sales_series = [round(val / 1000, 2) for val in perf_income_series]

    # Conversion funnel
    cart_count = Cart.objects.count()
    checkout_count = Order.objects.count()
    purchased_count = paid_orders_qs.count()
    impressions_count = max(cart_count * 8, total_users * 5, 1)
    conversion_rate = round((purchased_count / impressions_count) * 100, 2) if impressions_count else 0

    # Order statistics (weekly percentage)
    weekly_orders = Order.objects.filter(created_at__date__gte=last_7_days).count()
    order_statistics_percent = round((weekly_orders / total_orders) * 100, 1) if total_orders else 0

    # Finance tab series (last 7 months)
    finance_months = [((current_month_start - timedelta(days=1)).replace(day=1) - timedelta(days=30 * i)) for i in range(6, -1, -1)]
    finance_labels = [calendar.month_abbr[m.month] for m in finance_months]
    finance_income_series = []
    finance_expense_series = []
    finance_profit_series = []
    for m in finance_months:
        next_m = (m + timedelta(days=32)).replace(day=1)
        total = paid_orders_qs.filter(created_at__date__gte=m, created_at__date__lt=next_m).aggregate(total=Sum('total_amount'))['total'] or 0
        finance_income_series.append(float(total))
        finance_expense_series.append(float(total * Decimal('0.72')))
        finance_profit_series.append(float(total * Decimal('0.28')))

    weekly_expense = expenses_last_30 / 4 if expenses_last_30 else 0
    weekly_expense_percent = round((weekly_expense / expenses_last_30) * 100, 1) if expenses_last_30 else 0

    # Transactions list (recent orders)
    recent_transactions = []
    for order in recent_orders[:6]:
        recent_transactions.append({
            'method': order.payment_method,
            'label': order.user.get_full_name() or order.user.username,
            'amount': order.total_amount,
            'status': order.payment_status
        })

    # Top selling product (last 30 days)
    top_selling_item = (
        OrderItem.objects
        .filter(order__payment_status='PAID', order__created_at__date__gte=last_30_days)
        .values('product_id', 'product_name')
        .annotate(total_qty=Sum('quantity'), total_value=Sum('subtotal'))
        .order_by('-total_qty', '-total_value')
        .first()
    )
    top_selling_name = top_selling_item['product_name'] if top_selling_item else 'N/A'
    top_selling_sales = top_selling_item['total_value'] if top_selling_item else 0
    top_selling_qty = top_selling_item['total_qty'] if top_selling_item else 0
    top_selling_target_percent = min(round((top_selling_sales / (monthly_revenue or Decimal('1'))) * 100, 1), 100) if monthly_revenue else 0

    # Revenue growth vs previous month
    revenue_growth_percent = round(((monthly_revenue - prev_month_revenue) / prev_month_revenue) * 100, 1) if prev_month_revenue else (100.0 if monthly_revenue else 0.0)

    # Sales target
    sales_target = monthly_revenue * Decimal('1.2') if monthly_revenue else Decimal('10000')
    sales_target_percent = min(round((monthly_revenue / sales_target) * 100, 1), 100) if sales_target else 0
    
    # Recent Customers
    recent_customers = User.objects.filter(is_staff=False).order_by('-date_joined')[:5]
    
    context = {
        # Basic Stats
        'total_products': total_products,
        'active_products': active_products,
        'total_users': total_users,
        'total_orders': total_orders,
        
        # Revenue Stats
        'total_revenue': total_revenue,
        'monthly_revenue': monthly_revenue,
        'weekly_revenue': weekly_revenue,
        
        # Order Stats
        'pending_orders': pending_orders,
        'processing_orders': processing_orders,
        'shipped_orders': shipped_orders,
        'delivered_orders': delivered_orders,
        
        # Lists
        'recent_orders': recent_orders,
        'recent_order_items': recent_order_items,
        'top_products': top_products,
        'low_stock_products': low_stock_products,
        'recent_reviews': recent_reviews,
        'recent_customers': recent_customers,
        
        # Chart Data
        'daily_sales': list(daily_sales),

        # Dashboard UI data
        'top_selling_name': top_selling_name,
        'top_selling_sales': top_selling_sales,
        'top_selling_qty': top_selling_qty,
        'top_selling_target_percent': top_selling_target_percent,

        'visitors_weekly_percent': visitors_weekly_percent,
        'visitors_weekly_series': visitors_weekly_series,
        'visitors_today_series': visitors_today_series,
        'visitors_yesterday_series': visitors_yesterday_series,
        'visitors_last_month_series': visitors_last_month_series,
        'visitors_today_count': visitors_today_count,
        'visitors_yesterday_count': visitors_yesterday_count,
        'visitors_last_month_count': visitors_last_month_count,
        'activity_weekly_percent': activity_weekly_percent,
        'activity_weekly_series': activity_weekly_series,

        'last_30_revenue': last_30_revenue,
        'sales_today': sales_today,
        'sales_yesterday': sales_yesterday,
        'sales_last_month': sales_last_month,
        'sales_last_6_months': sales_last_6_months,
        'sales_change_percent': sales_change_percent,
        'profit_last_30': profit_last_30,
        'expenses_last_30': expenses_last_30,
        'expenses_percent': expenses_percent,
        'expenses_remaining_percent': expenses_remaining_percent,
        'transactions_last_30': transactions_last_30,
        'transactions_change_percent': transactions_change_percent,

        'month_labels': month_labels,
        'income_yearly_series': income_yearly_series,
        'expense_yearly_series': expense_yearly_series,

        'report_income': report_income,
        'report_expense': report_expense,
        'report_profit': report_profit,
        'report_income_change': report_income_change,

        'perf_month_labels': perf_month_labels,
        'perf_income_series': perf_income_series,
        'perf_sales_series': perf_sales_series,

        'impressions_count': impressions_count,
        'cart_count': cart_count,
        'checkout_count': checkout_count,
        'purchased_count': purchased_count,
        'conversion_rate': conversion_rate,

        'order_statistics_percent': order_statistics_percent,
        'weekly_orders': weekly_orders,
        'total_sales_count': purchased_count,

        'finance_labels': finance_labels,
        'finance_income_series': finance_income_series,
        'finance_expense_series': finance_expense_series,
        'finance_profit_series': finance_profit_series,
        'weekly_expense': weekly_expense,
        'weekly_expense_percent': weekly_expense_percent,

        'recent_transactions': recent_transactions,

        'revenue_growth_percent': revenue_growth_percent,
        'sales_target': sales_target,
        'sales_target_percent': sales_target_percent,
    }
    return render(request, 'admin_panel/dashboard.html', context)

@login_required(login_url='login')
@staff_member_required(login_url='login')
def admin_add_product(request):
    """Admin Add Product Page"""
    if request.method == 'POST':
        try:
            # Get basic form data
            name = request.POST.get('name')
            price = request.POST.get('price')
            old_price = request.POST.get('old_price')
            stock = request.POST.get('stock')
            is_active = request.POST.get('is_active') == 'on'
            is_top_deal = request.POST.get('is_top_deal') == 'on'
            
            # Convert rating and review_count to proper types
            try:
                rating = float(request.POST.get('rating', 0)) or 0
            except (TypeError, ValueError):
                rating = 0
            
            try:
                review_count = int(float(request.POST.get('review_count', 0))) or 0
            except (TypeError, ValueError):
                review_count = 0
            
            # Get new fields
            category = request.POST.get('category')
            sku = request.POST.get('sku', '')
            brand = request.POST.get('brand', '')
            description = request.POST.get('description', '')
            weight = request.POST.get('weight', '')
            color = request.POST.get('color', '')
            # Handle multiple size selections from checkboxes
            size_list = request.POST.getlist('size')
            size = ', '.join(size_list) if size_list else ''
            
            # Get images
            image = request.FILES.get('image')
            descriptionImage = request.FILES.get('descriptionImage')
            gallery_images = request.FILES.getlist('gallery_images')
            
            # Create product
            product = Product.objects.create(
                name=name,
                price=price,
                old_price=old_price if old_price else None,
                stock=stock,
                rating=rating,
                review_count=review_count,
                is_active=True,
                is_top_deal=False,
                image=image,
                descriptionImage=descriptionImage,
                category=category if category else None,
                sku=sku,
                brand=brand,
                description=description,
                weight=weight,
                color=color,
                size=size
            )
            
            # Add gallery images if provided
            for idx, gallery_image in enumerate(gallery_images, start=1):
                ProductImage.objects.create(
                    product=product,
                    image=gallery_image,
                    order=idx,
                    is_active=True
                )

            # Auto-generate approved reviews if rating/review_count provided
            if review_count > 0 and rating > 0:
                generate_auto_reviews(product, review_count, rating, request.user)
            
            messages.success(request, f'Product "{product.name}" added successfully with {len(gallery_images)} gallery images!')
            return redirect('admin_add_product')
            
        except Exception as e:
            messages.error(request, f'Error adding product: {str(e)}')
    
    # Get categories for dropdown
    categories = CategoryIcon.objects.filter(is_active=True).order_by('order', 'id')
    return render(request, 'admin_panel/add_product.html', {'categories': categories})


def generate_auto_reviews(product, review_total, target_avg, reviewer_user=None):
    """Generate approved reviews and rating distribution for a product."""
    try:
        review_total = int(review_total)
    except (TypeError, ValueError):
        return

    try:
        target_avg = float(target_avg)
    except (TypeError, ValueError):
        return

    if review_total <= 0 or target_avg <= 0:
        return

    target_avg = max(1.0, min(5.0, target_avg))

    # Gaussian-like weights around average
    weights = {}
    for r in range(1, 6):
        dist = (r - target_avg)
        weights[r] = pow(2.718281828, -(dist ** 2) / 2)

    weight_sum = sum(weights.values()) or 1
    raw = {r: (weights[r] / weight_sum) * review_total for r in range(1, 6)}

    counts = {r: int(raw[r]) for r in range(1, 6)}
    remainder = review_total - sum(counts.values())
    if remainder > 0:
        remainders = sorted(
            [(r, raw[r] - counts[r]) for r in range(1, 6)],
            key=lambda x: x[1],
            reverse=True
        )
        for r, _ in remainders[:remainder]:
            counts[r] += 1

    target_sum = int(round(target_avg * review_total))
    current_sum = sum(r * c for r, c in counts.items())

    while current_sum < target_sum:
        moved = False
        for r in range(4, 0, -1):
            if counts[r] > 0:
                counts[r] -= 1
                counts[r + 1] += 1
                current_sum += 1
                moved = True
                if current_sum >= target_sum:
                    break
        if not moved:
            break

    while current_sum > target_sum:
        moved = False
        for r in range(2, 6):
            if counts[r] > 0:
                counts[r] -= 1
                counts[r - 1] += 1
                current_sum -= 1
                moved = True
                if current_sum <= target_sum:
                    break
        if not moved:
            break

    if reviewer_user is None:
        reviewer_user = (
            User.objects.filter(is_superuser=True).first()
            or User.objects.filter(is_staff=True).first()
            or User.objects.first()
        )

    if reviewer_user is None:
        return

    reviewer_name = "Admin"
    reviewer_email = "admin@fashionhub.local"
    if reviewer_user:
        reviewer_name = reviewer_user.get_full_name().strip() or reviewer_user.username
        reviewer_email = reviewer_user.email or reviewer_email

    for r in range(5, 0, -1):
        for _ in range(counts.get(r, 0)):
            ProductReview.objects.create(
                product=product,
                user=reviewer_user,
                rating=r,
                name=reviewer_name,
                email=reviewer_email,
                comment="",
                is_approved=True,
                is_auto_generated=True
            )

@login_required(login_url='login')
@staff_member_required(login_url='login')
def admin_product_list(request):
    """Admin Product List Page"""
    products = Product.objects.all().order_by('-id')

    # Filters
    status_filter = request.GET.get('status', 'all')
    category_filter = request.GET.get('category', 'all')
    stock_filter = request.GET.get('stock', 'all')
    search_query = request.GET.get('q', '').strip()

    if status_filter == 'active':
        products = products.filter(is_active=True)
    elif status_filter == 'inactive':
        products = products.filter(is_active=False)

    if stock_filter == 'in':
        products = products.filter(stock__gt=0)
    elif stock_filter == 'out':
        products = products.filter(stock__lte=0)
    elif stock_filter == 'low':
        products = products.filter(stock__gt=0, stock__lte=10)

    if search_query:
        products = products.filter(
            Q(name__icontains=search_query) |
            Q(sku__icontains=search_query) |
            Q(brand__icontains=search_query)
        )

    # Export CSV
    if request.GET.get('export') == '1':
        import csv
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="products.csv"'
        writer = csv.writer(response)
        writer.writerow(['ID', 'Name', 'Category', 'Stock', 'SKU', 'Price', 'Status'])
        for product in products:
            writer.writerow([
                product.id,
                product.name,
                'General',
                product.stock,
                product.sku or '',
                product.price,
                'Active' if product.is_active else 'Inactive'
            ])
        return response
    
    # Pagination
    per_page = request.GET.get('per_page', '10')
    try:
        per_page = int(per_page)
    except ValueError:
        per_page = 10
    per_page = per_page if per_page in [10, 20, 50, 100] else 10

    paginator = Paginator(products, per_page)
    page = request.GET.get('page')
    
    try:
        products = paginator.page(page)
    except PageNotAnInteger:
        products = paginator.page(1)
    except EmptyPage:
        products = paginator.page(paginator.num_pages)

    # Summary cards
    paid_orders = Order.objects.filter(payment_status='PAID')
    in_store_orders = paid_orders.filter(payment_method='COD')
    website_orders = paid_orders.exclude(payment_method='COD')

    in_store_sales = in_store_orders.aggregate(total=Sum('total_amount'))['total'] or 0
    website_sales = website_orders.aggregate(total=Sum('total_amount'))['total'] or 0
    discount_totals = paid_orders.aggregate(subtotal=Sum('subtotal'), total=Sum('total_amount'))
    discount_total = (discount_totals['subtotal'] or 0) - (discount_totals['total'] or 0)
    affiliate_sales = paid_orders.filter(is_resell=True).aggregate(total=Sum('total_amount'))['total'] or 0

    in_store_orders_count = in_store_orders.count()
    website_orders_count = website_orders.count()
    discount_orders_count = paid_orders.filter(subtotal__gt=0).count()
    affiliate_orders_count = paid_orders.filter(is_resell=True).count()

    query_params = request.GET.copy()
    query_params.pop('page', None)
    query_params.pop('export', None)

    context = {
        'products': products,
        'status_filter': status_filter,
        'category_filter': category_filter,
        'stock_filter': stock_filter,
        'search_query': search_query,
        'per_page': per_page,

        'in_store_sales': in_store_sales,
        'website_sales': website_sales,
        'discount_total': discount_total,
        'affiliate_sales': affiliate_sales,
        'in_store_orders_count': in_store_orders_count,
        'website_orders_count': website_orders_count,
        'discount_orders_count': discount_orders_count,
        'affiliate_orders_count': affiliate_orders_count,
        'querystring': query_params.urlencode(),
    }
    return render(request, 'admin_panel/product_list.html', context)

@login_required(login_url='login')
@staff_member_required(login_url='login')
def admin_edit_product(request, product_id):
    """Admin Edit Product Page"""
    product = get_object_or_404(Product, id=product_id)
    
    if request.method == 'POST':
        try:
            # Update basic fields
            product.name = request.POST.get('name')
            product.price = request.POST.get('price')
            product.old_price = request.POST.get('old_price') or None
            product.stock = request.POST.get('stock')
            product.category = request.POST.get('category') or None
            product.rating = request.POST.get('rating', 0)
            product.review_count = request.POST.get('review_count', 0)
            product.is_active = request.POST.get('is_active') == 'on'
            product.is_top_deal = request.POST.get('is_top_deal') == 'on'
            
            # Update new fields
            product.sku = request.POST.get('sku', '')
            product.brand = request.POST.get('brand', '')
            product.tags = request.POST.get('tags', '')
            product.description = request.POST.get('description', '')
            product.weight = request.POST.get('weight', '')
            product.dimensions = request.POST.get('dimensions', '')
            product.color = request.POST.get('color', '')
            size_list = request.POST.getlist('size')
            product.size = ', '.join(size_list) if size_list else ''
            product.shipping_info = request.POST.get('shipping_info', '')
            product.care_info = request.POST.get('care_info', '')
            
            # Update image if provided
            if 'image' in request.FILES:
                product.image = request.FILES['image']
            
            # Update description image if provided
            if 'descriptionImage' in request.FILES:
                product.descriptionImage = request.FILES['descriptionImage']
            
            product.save()
            
            # Handle gallery images if provided
            if 'gallery_images' in request.FILES:
                gallery_images = request.FILES.getlist('gallery_images')
                for idx, gallery_image in enumerate(gallery_images, start=ProductImage.objects.filter(product=product).count() + 1):
                    ProductImage.objects.create(
                        product=product,
                        image=gallery_image,
                        order=idx,
                        is_active=True
                    )
            
            messages.success(request, f'Product "{product.name}" updated successfully!')
            return redirect('admin_product_list')
        
        except Exception as e:
            messages.error(request, f'Error updating product: {str(e)}')
            return redirect('admin_edit_product', product_id=product_id)
    
    context = {
        'product': product,
        'categories': CategoryIcon.objects.filter(is_active=True).order_by('order', 'id'),
    }
    return render(request, 'admin_panel/edit_product.html', context)

@login_required(login_url='login')
@staff_member_required(login_url='login')
def admin_delete_product(request, product_id):
    """Delete Product"""
    product = get_object_or_404(Product, id=product_id)
    product_name = product.name
    product.delete()
    
    messages.success(request, f'Product "{product_name}" deleted successfully!')
    return redirect('admin_product_list')

@login_required(login_url='login')
@staff_member_required(login_url='login')
def admin_categories(request):
    """Admin Category Management Page"""
    categories = CategoryIcon.objects.all().order_by('order', 'id')
    
    context = {
        'categories': categories,
    }
    return render(request, 'admin_panel/categories.html', context)

@login_required(login_url='login')
@staff_member_required(login_url='login')
def admin_add_category(request):
    """Admin Add Category Page"""
    if request.method == 'POST':
        try:
            name = request.POST.get('name')
            category_key = request.POST.get('category_key')
            icon_class = request.POST.get('icon_class', '')
            icon_color = request.POST.get('icon_color', '#0288d1')
            background_gradient = request.POST.get('background_gradient', 'linear-gradient(135deg, #e0f7ff 0%, #b3e5fc 100%)')
            order = request.POST.get('order', 0)
            is_active = request.POST.get('is_active') == 'on'
            
            category = CategoryIcon.objects.create(
                name=name,
                category_key=category_key,
                icon_class=icon_class,
                icon_color=icon_color,
                background_gradient=background_gradient,
                order=order,
                is_active=is_active
            )
            
            # Handle icon image upload
            if request.FILES.get('icon_image'):
                category.icon_image = request.FILES['icon_image']
                category.save()
            
            messages.success(request, f'Category "{category.name}" added successfully!')
            return redirect('admin_categories')
        
        except Exception as e:
            messages.error(request, f'Error adding category: {str(e)}')
            return redirect('admin_add_category')
    
    return render(request, 'admin_panel/add_category.html')

@login_required(login_url='login')
@staff_member_required(login_url='login')
def admin_edit_category(request, category_id):
    """Admin Edit Category Page"""
    category = get_object_or_404(CategoryIcon, id=category_id)
    
    if request.method == 'POST':
        try:
            category.name = request.POST.get('name')
            category.category_key = request.POST.get('category_key')
            category.icon_class = request.POST.get('icon_class', '')
            category.icon_color = request.POST.get('icon_color', '#0288d1')
            category.background_gradient = request.POST.get('background_gradient', 'linear-gradient(135deg, #e0f7ff 0%, #b3e5fc 100%)')
            category.order = request.POST.get('order', 0)
            category.is_active = request.POST.get('is_active') == 'on'
            
            # Handle icon image upload
            if request.FILES.get('icon_image'):
                category.icon_image = request.FILES['icon_image']
            
            category.save()
            
            messages.success(request, f'Category "{category.name}" updated successfully!')
            return redirect('admin_categories')
        
        except Exception as e:
            messages.error(request, f'Error updating category: {str(e)}')
            return redirect('admin_edit_category', category_id=category_id)
    
    context = {
        'category': category,
    }
    return render(request, 'admin_panel/edit_category.html', context)

@login_required(login_url='login')
@staff_member_required(login_url='login')
def admin_delete_category(request, category_id):
    """Delete Category"""
    category = get_object_or_404(CategoryIcon, id=category_id)
    category_name = category.name
    category.delete()
    
    messages.success(request, f'Category "{category_name}" deleted successfully!')
    return redirect('admin_categories')

@login_required(login_url='login')
@staff_member_required(login_url='login')
def admin_brand_partners(request):
    """Admin Brand Partners Management Page"""
    brand_partners = BrandPartner.objects.all().order_by('order', 'name')
    
    context = {
        'brand_partners': brand_partners,
    }
    return render(request, 'admin_panel/brand_partners.html', context)

@login_required(login_url='login')
@staff_member_required(login_url='login')
def admin_add_brand_partner(request):
    """Add New Brand Partner"""
    if request.method == 'POST':
        name = request.POST.get('name')
        logo = request.FILES.get('logo')
        link_url = request.POST.get('link_url', '')
        order = request.POST.get('order', 0)
        is_active = request.POST.get('is_active') == 'on'
        
        if name and logo:
            partner = BrandPartner.objects.create(
                name=name,
                logo=logo,
                link_url=link_url,
                order=order,
                is_active=is_active
            )
            messages.success(request, f'Brand Partner "{name}" added successfully!')
            return redirect('admin_brand_partners')
        else:
            messages.error(request, 'Please provide brand name and logo!')
    
    return render(request, 'admin_panel/add_brand_partner.html')

@login_required(login_url='login')
@staff_member_required(login_url='login')
def admin_edit_brand_partner(request, partner_id):
    """Edit Brand Partner"""
    partner = get_object_or_404(BrandPartner, id=partner_id)
    
    if request.method == 'POST':
        partner.name = request.POST.get('name')
        link_url = request.POST.get('link_url', '')
        partner.link_url = link_url
        partner.order = request.POST.get('order', 0)
        partner.is_active = request.POST.get('is_active') == 'on'
        
        # Update logo if new one uploaded
        if request.FILES.get('logo'):
            partner.logo = request.FILES.get('logo')
        
        partner.save()
        messages.success(request, f'Brand Partner "{partner.name}" updated successfully!')
        return redirect('admin_brand_partners')
    
    context = {
        'partner': partner,
    }
    return render(request, 'admin_panel/edit_brand_partner.html', context)


@login_required(login_url='login')
@staff_member_required(login_url='login')
def admin_site_settings(request):
    """Admin Site Settings - Logo, Name, Contact Info"""
    settings_obj = SiteSettings.get_settings()
    
    if request.method == 'POST':
        settings_obj.site_name = request.POST.get('site_name', 'VibeMall')
        settings_obj.site_name_html = request.POST.get('site_name_html', '')
        settings_obj.tagline = request.POST.get('tagline', '')
        settings_obj.contact_email = request.POST.get('contact_email', 'support@vibemall.com')
        settings_obj.contact_phone = request.POST.get('contact_phone', '+91 1234567890')
        settings_obj.facebook_url = request.POST.get('facebook_url', '')
        settings_obj.instagram_url = request.POST.get('instagram_url', '')
        settings_obj.twitter_url = request.POST.get('twitter_url', '')
        settings_obj.youtube_url = request.POST.get('youtube_url', '')
        
        # Handle logo uploads
        if request.FILES.get('site_logo'):
            settings_obj.site_logo = request.FILES.get('site_logo')
        
        if request.FILES.get('site_favicon'):
            settings_obj.site_favicon = request.FILES.get('site_favicon')
        
        if request.FILES.get('admin_logo'):
            settings_obj.admin_logo = request.FILES.get('admin_logo')
        
        settings_obj.save()
        messages.success(request, 'Site settings updated successfully!')
        return redirect('admin_site_settings')
    
    context = {
        'settings': settings_obj,
    }
    return render(request, 'admin_panel/site_settings.html', context)


@login_required(login_url='login')
@staff_member_required(login_url='login')
def admin_delete_brand_partner(request, partner_id):
    """Delete Brand Partner"""
    partner = get_object_or_404(BrandPartner, id=partner_id)
    partner_name = partner.name
    partner.delete()
    
    messages.success(request, f'Brand Partner "{partner_name}" deleted successfully!')
    return redirect('admin_brand_partners')

@login_required(login_url='login')
@staff_member_required(login_url='login')
def admin_reviews(request):
    """Admin Review Management Page"""
    # Handle bulk actions
    if request.method == 'POST':
        action = request.POST.get('action')
        review_ids = request.POST.getlist('review_ids')
        
        if action == 'approve' and review_ids:
            ProductReview.objects.filter(id__in=review_ids).update(is_approved=True)
            messages.success(request, f'{len(review_ids)} review(s) approved successfully!')
        elif action == 'reject' and review_ids:
            ProductReview.objects.filter(id__in=review_ids).update(is_approved=False)
            messages.success(request, f'{len(review_ids)} review(s) rejected successfully!')
        elif action == 'delete' and review_ids:
            ProductReview.objects.filter(id__in=review_ids).delete()
            messages.success(request, f'{len(review_ids)} review(s) deleted successfully!')
        
        return redirect('admin_reviews')
    
    # Filter reviews
    filter_status = request.GET.get('status', 'all')
    reviews = ProductReview.objects.all().select_related('product', 'user').order_by('-created_at')
    
    if filter_status == 'pending':
        reviews = reviews.filter(is_approved=False)
    elif filter_status == 'approved':
        reviews = reviews.filter(is_approved=True)
    
    # Count stats
    total_reviews = ProductReview.objects.count()
    pending_reviews = ProductReview.objects.filter(is_approved=False).count()
    approved_reviews = ProductReview.objects.filter(is_approved=True).count()
    
    # Pagination
    paginator = Paginator(reviews, 20)
    page = request.GET.get('page')
    
    try:
        reviews = paginator.page(page)
    except PageNotAnInteger:
        reviews = paginator.page(1)
    except EmptyPage:
        reviews = paginator.page(paginator.num_pages)
    
    context = {
        'reviews': reviews,
        'filter_status': filter_status,
        'total_reviews': total_reviews,
        'pending_reviews': pending_reviews,
        'approved_reviews': approved_reviews,
    }
    return render(request, 'admin_panel/reviews.html', context)


# ===== ORDER MANAGEMENT (ADMIN PANEL) =====

# Helper Functions for Order Management
def send_order_status_email(order):
    """Send email notification to customer when order status changes"""
    from django.core.mail import send_mail
    from django.template.loader import render_to_string
    
    try:
        subject = f'Order {order.order_number} - Status Update'
        message = render_to_string('emails/order_status_update.html', {
            'order': order,
            'customer_name': order.user.get_full_name() or order.user.username
        })
        send_mail(
            subject,
            '',
            settings.DEFAULT_FROM_EMAIL,
            [order.user.email],
            html_message=message,
            fail_silently=True
        )
    except Exception as e:
        print(f"Email sending failed: {e}")


def send_admin_new_order_notification(order):
    """Send email notification to admin when new order is received"""
    from django.core.mail import send_mail
    from django.template.loader import render_to_string
    
    try:
        # Get admin email from settings (configurable)
        admin_settings = AdminEmailSettings.objects.first()
        admin_email = admin_settings.admin_email if admin_settings and admin_settings.is_active else 'rajpaladiya2023@gmail.com'
        
        subject = f'New Order Received - {order.order_number}'
        message = f"""
        New Order Received!
        
        Order Number: {order.order_number}
        Customer: {order.user.get_full_name() or order.user.username}
        Email: {order.user.email}
        Phone: {order.user.userprofile.mobile_number if hasattr(order.user, 'userprofile') else 'N/A'}
        
        Order Details:
        Total Amount: ₹{order.total_amount}
        Payment Method: {order.get_payment_method_display()}
        Payment Status: {order.get_payment_status_display()}
        Order Status: {order.get_order_status_display()}
        
        Items:
        """
        
        for item in order.items.all():
            message += f"\n- {item.product_name} x {item.quantity} = ₹{item.subtotal}"
        
        message += f"\n\nShipping Address:\n{order.shipping_address}"
        message += f"\n\nOrder Date: {order.created_at.strftime('%d %b %Y, %I:%M %p')}"
        message += f"\n\nView Order: http://localhost:8000/admin-panel/orders/{order.id}/"
        
        send_mail(
            subject,
            message,
            settings.DEFAULT_FROM_EMAIL,
            [admin_email],
            fail_silently=True
        )
    except Exception as e:
        print(f"Admin email notification failed: {e}")


def export_orders_to_excel(orders):
    """Export orders to Excel file"""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    from django.http import HttpResponse
    from datetime import datetime
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Orders"
    
    # Headers
    headers = ['Order Number', 'Product', 'Quantity', 'Price', 'Customer Name', 'Email', 'Phone', 
               'Order Status', 'Payment Status', 'Order Date', 'Tracking Number']
    ws.append(headers)
    
    # Style headers
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    # Data rows
    for item in orders:
        ws.append([
            item.order.order_number,
            item.product_name,
            item.quantity,
            float(item.subtotal),
            item.order.user.get_full_name() or item.order.user.username,
            item.order.user.email,
            item.order.user.userprofile.mobile_number if hasattr(item.order.user, 'userprofile') else '',
            item.order.get_order_status_display(),
            item.order.get_payment_status_display(),
            item.order.created_at.strftime('%d-%m-%Y'),
            item.order.tracking_number or ''
        ])
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Create response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="Order {datetime.now().strftime("%m/%Y")}.xlsx"'
    wb.save(response)
    
    return response


@login_required(login_url='login')
@staff_member_required(login_url='login')
def admin_orders(request):
    """Comprehensive Admin Order Management with 24+ features"""
    from django.core.paginator import Paginator
    from django.http import HttpResponse, JsonResponse
    from django.db.models import Sum, Count, Q
    import csv
    from datetime import datetime, timedelta
    
    # Handle Bulk Actions (POST)
    if request.method == 'POST':
        action = request.POST.get('action')
        order_ids = request.POST.getlist('order_ids')
        
        if action and order_ids:
            orders = Order.objects.filter(id__in=order_ids)
            
            if action in ['PENDING', 'PROCESSING', 'SHIPPED', 'DELIVERED', 'CANCELLED']:
                # Bulk status update
                from django.utils import timezone
                for order in orders:
                    old_status = order.order_status
                    order.order_status = action
                    # Set delivery_date and payment_status when status changes to DELIVERED
                    if action == 'DELIVERED' and old_status != 'DELIVERED':
                        order.delivery_date = timezone.now()
                        order.payment_status = 'PAID'
                    order.save()
                    # Create status history
                    OrderStatusHistory.objects.create(
                        order=order,
                        old_status=old_status,
                        new_status=action,
                        changed_by=request.user,
                        notes=f"Bulk status update by {request.user.username}"
                    )
                    # Send email notification to customer with beautiful template
                    send_order_status_update_email(order, old_status, action)
                messages.success(request, f"{len(order_ids)} orders updated to {action}")
                
            elif action == 'delete':
                # Bulk delete
                orders.delete()
                messages.success(request, f"{len(order_ids)} orders deleted")
                
            return redirect('admin_orders')
    
    # Get all order items (individual products with their order details)
    all_orders = OrderItem.objects.select_related('order', 'order__user', 'product').order_by('-order__created_at')
    
    # === FILTERS ===
    # Status filter
    status_filter = request.GET.get('status', '')
    if status_filter:
        all_orders = all_orders.filter(order__order_status=status_filter)
    
    # Payment status filter
    payment_filter = request.GET.get('payment', '')
    if payment_filter:
        all_orders = all_orders.filter(order__payment_status=payment_filter)

    # Payment method filter
    payment_method_filter = request.GET.get('payment_method', '')
    if payment_method_filter:
        all_orders = all_orders.filter(order__payment_method=payment_method_filter)
    
    # Approval status filter
    approval_filter = request.GET.get('approval', '')
    if approval_filter:
        all_orders = all_orders.filter(order__approval_status=approval_filter)
    
    # Suspicious orders filter
    suspicious_filter = request.GET.get('suspicious', '')
    if suspicious_filter == '1':
        all_orders = all_orders.filter(order__is_suspicious=True)
    elif suspicious_filter == '0':
        all_orders = all_orders.filter(order__is_suspicious=False)
    
    # Date range filter
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    if date_from:
        all_orders = all_orders.filter(order__created_at__gte=date_from)
    if date_to:
        all_orders = all_orders.filter(order__created_at__lte=date_to)
    
    # Search filter (Order ID, Customer name, Phone, Email, Product name)
    search_query = request.GET.get('search', '')
    if search_query:
        all_orders = all_orders.filter(
            Q(order__order_number__icontains=search_query) |
            Q(order__user__username__icontains=search_query) |
            Q(order__user__email__icontains=search_query) |
            Q(order__user__first_name__icontains=search_query) |
            Q(order__user__last_name__icontains=search_query) |
            Q(order__user__userprofile__mobile_number__icontains=search_query) |
            Q(order__tracking_number__icontains=search_query) |
            Q(product_name__icontains=search_query)
        )
    
    # Sorting
    sort_by = request.GET.get('sort', '-order__created_at')
    # Fix sorting for fields that belong to Order model (since we're querying OrderItem)
    if sort_by == 'total_amount':
        sort_by = 'order__total_amount'
    elif sort_by == '-total_amount':
        sort_by = '-order__total_amount'
    elif sort_by == 'order_status':
        sort_by = 'order__order_status'
    elif sort_by == '-order_status':
        sort_by = '-order__order_status'
    elif sort_by == 'payment_status':
        sort_by = 'order__payment_status'
    elif sort_by == '-payment_status':
        sort_by = '-order__payment_status'
    all_orders = all_orders.order_by(sort_by)
    
    # === STATISTICS ===
    total_orders = Order.objects.count()
    pending_orders = Order.objects.filter(order_status='PENDING').count()
    processing_orders = Order.objects.filter(order_status='PROCESSING').count()
    shipped_orders = Order.objects.filter(order_status='SHIPPED').count()
    delivered_orders = Order.objects.filter(order_status='DELIVERED').count()
    cancelled_orders = Order.objects.filter(order_status='CANCELLED').count()
    
    # Approval stats
    pending_approval_orders = Order.objects.filter(approval_status='PENDING_APPROVAL').count()
    approved_orders = Order.objects.filter(approval_status='APPROVED').count()
    rejected_orders = Order.objects.filter(approval_status='REJECTED').count()
    suspicious_orders = Order.objects.filter(is_suspicious=True).count()
    
    # Revenue stats (use Order model, not OrderItem)
    from decimal import Decimal
    paid_orders = Order.objects.filter(payment_status='PAID')
    total_revenue = paid_orders.aggregate(total=Sum('total_amount'))['total'] or Decimal('0')
    today_revenue = paid_orders.filter(created_at__date=datetime.now().date()).aggregate(total=Sum('total_amount'))['total'] or Decimal('0')
    this_month_revenue = paid_orders.filter(created_at__month=datetime.now().month).aggregate(total=Sum('total_amount'))['total'] or Decimal('0')
    
    # Profit calculation (assuming 30% profit margin)
    profit_margin = Decimal('0.30')
    total_profit = total_revenue * profit_margin
    today_profit = today_revenue * profit_margin
    this_month_profit = this_month_revenue * profit_margin
    
    # Today's orders
    today_orders = Order.objects.filter(created_at__date=datetime.now().date()).count()
    
    # Payment methods list
    payment_methods = list(
        Order.objects.exclude(payment_method__isnull=True)
        .exclude(payment_method__exact='')
        .values_list('payment_method', flat=True)
        .distinct()
    )

    # === EXPORT TO CSV ===
    if request.GET.get('export') == 'csv':
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="Order {datetime.now().strftime("%m/%Y")}.csv"'
        
        writer = csv.writer(response)
        writer.writerow(['Order Number', 'Product', 'Quantity', 'Price', 'Customer', 'Email', 'Phone', 'Order Status', 'Payment Status', 'Date'])
        
        for item in all_orders:
            writer.writerow([
                item.order.order_number,
                item.product_name,
                item.quantity,
                item.subtotal,
                item.order.user.get_full_name() or item.order.user.username,
                item.order.user.email,
                item.order.user.userprofile.mobile_number if hasattr(item.order.user, 'userprofile') else '',
                item.order.get_order_status_display(),
                item.order.get_payment_status_display(),
                item.order.created_at.strftime('%d-%m-%Y')
            ])
        
        return response
    
    # === EXPORT TO EXCEL ===
    if request.GET.get('export') == 'excel':
        return export_orders_to_excel(all_orders)
    
    # === PAGINATION ===
    page_size = request.GET.get('page_size', '20')
    paginator = Paginator(all_orders, int(page_size))
    page_number = request.GET.get('page', 1)
    orders = paginator.get_page(page_number)
    
    context = {
        'orders': orders,
        'status_filter': status_filter,
        'payment_filter': payment_filter,
        'payment_method_filter': payment_method_filter,
        'approval_filter': approval_filter,
        'suspicious_filter': suspicious_filter,
        'search_query': search_query,
        'date_from': date_from,
        'date_to': date_to,
        'sort_by': sort_by,
        'page_size': page_size,
        
        # Statistics
        'total_orders': total_orders,
        'pending_orders': pending_orders,
        'processing_orders': processing_orders,
        'shipped_orders': shipped_orders,
        'delivered_orders': delivered_orders,
        'cancelled_orders': cancelled_orders,
        
        # Approval stats
        'pending_approval_orders': pending_approval_orders,
        'approved_orders': approved_orders,
        'rejected_orders': rejected_orders,
        'suspicious_orders': suspicious_orders,
        
        # Revenue
        'total_revenue': total_revenue,
        'today_revenue': today_revenue,
        'this_month_revenue': this_month_revenue,
        'today_orders': today_orders,

        # Payment methods
        'payment_methods': payment_methods,
        
        # Profit
        'total_profit': total_profit,
        'today_profit': today_profit,
        'this_month_profit': this_month_profit,
    }
    
    return render(request, 'admin_panel/orders.html', context)


@login_required(login_url='login')
@staff_member_required(login_url='login')
def admin_order_details(request, order_id):
    """Order Details with timeline, tracking, notes, and actions"""
    order = get_object_or_404(Order, id=order_id)
    
    # Handle POST actions
    if request.method == 'POST':
        action = request.POST.get('action')
        
        if action == 'update_status':
            from django.utils import timezone
            old_status = order.order_status
            new_status = request.POST.get('new_status')
            order.order_status = new_status
            # Set delivery_date and payment_status when status changes to DELIVERED
            if new_status == 'DELIVERED' and old_status != 'DELIVERED':
                order.delivery_date = timezone.now()
                order.payment_status = 'PAID'
            order.save()
            
            # Create status history
            OrderStatusHistory.objects.create(
                order=order,
                old_status=old_status,
                new_status=new_status,
                changed_by=request.user,
                notes=request.POST.get('status_notes', '')
            )
            
            # Send email notification with beautiful template
            send_order_status_update_email(order, old_status, new_status)
            messages.success(request, f'Order status updated to {new_status}')
            
        elif action == 'add_tracking':
            order.tracking_number = request.POST.get('tracking_number')
            order.courier_name = request.POST.get('courier_name')
            order.save()
            messages.success(request, 'Tracking information updated')
            
        elif action == 'update_payment':
            old_payment_status = order.payment_status
            new_payment_status = request.POST.get('new_payment_status') or request.POST.get('payment_status')
            order.payment_status = new_payment_status
            order.save()
            messages.success(request, f'Payment status updated to {order.get_payment_status_display()}')
            
        elif action == 'add_note':
            order.admin_notes = request.POST.get('admin_notes')
            order.save()
            messages.success(request, 'Admin notes saved')
        
        return redirect('admin_order_details', order_id=order.id)
    
    # Get status history
    status_history = order.status_history.all()
    
    context = {
        'order': order,
        'status_history': status_history,
    }
    return render(request, 'admin_panel/order_details.html', context)


@login_required(login_url='login')
@staff_member_required(login_url='login')
def admin_api_search_orders(request):
    """Dynamic order search for admin navbar (order ID / number)."""
    from django.http import JsonResponse
    from django.db.models import Q
    from django.urls import reverse

    query = (request.GET.get('q') or '').strip()
    if not query:
        return JsonResponse({'results': []})

    orders = Order.objects.select_related('user').filter(
        Q(order_number__icontains=query)
    )

    if query.isdigit():
        orders = orders | Order.objects.select_related('user').filter(id=int(query))

    orders = orders.order_by('-created_at')[:10]

    results = []
    for order in orders:
        customer_name = f"{order.user.first_name} {order.user.last_name}".strip() or order.user.username
        results.append({
            'id': order.id,
            'order_number': order.order_number,
            'customer': customer_name,
            'total_amount': float(order.total_amount),
            'order_status': order.order_status,
            'payment_status': order.payment_status,
            'created_at': order.created_at.strftime('%d %b %Y, %I:%M %p'),
            'detail_url': reverse('admin_order_details', args=[order.id]),
        })

    return JsonResponse({'results': results})


@login_required(login_url='login')
@staff_member_required(login_url='login')
def admin_api_search_orders(request):
    """Dynamic order search for admin navbar (order ID / number)."""
    from django.http import JsonResponse
    from django.db.models import Q
    from django.urls import reverse

    query = (request.GET.get('q') or '').strip()
    if not query:
        return JsonResponse({'results': []})

    orders = Order.objects.select_related('user').filter(
        Q(order_number__icontains=query)
    )

    if query.isdigit():
        orders = orders | Order.objects.select_related('user').filter(id=int(query))

    orders = orders.order_by('-created_at')[:10]

    results = []
    for order in orders:
        customer_name = f"{order.user.first_name} {order.user.last_name}".strip() or order.user.username
        results.append({
            'id': order.id,
            'order_number': order.order_number,
            'customer': customer_name,
            'total_amount': float(order.total_amount),
            'order_status': order.order_status,
            'payment_status': order.payment_status,
            'created_at': order.created_at.strftime('%d %b %Y, %I:%M %p'),
            'detail_url': reverse('admin_order_details', args=[order.id]),
        })

    return JsonResponse({'results': results})


@login_required(login_url='login')
@staff_member_required(login_url='login')
def admin_approve_order(request, order_id):
    """Approve a pending order"""
    order = get_object_or_404(Order, id=order_id)
    
    if order.approval_status != 'PENDING_APPROVAL':
        messages.warning(request, 'Order is not pending approval.')
        return redirect(request.META.get('HTTP_REFERER', 'admin_orders'))
    
    # Update approval status
    from django.utils import timezone
    order.approval_status = 'APPROVED'
    order.approved_by = request.user
    order.approved_at = timezone.now()
    order.order_status = 'PROCESSING'
    
    # Add approval notes if provided
    if request.method == 'POST':
        notes = request.POST.get('approval_notes', '')
        if notes:
            order.approval_notes = notes
    else:
        order.approval_notes = f'Manually approved by {request.user.username}'
    
    order.save()
    
    # Send approval email to customer
    try:
        from django.template.loader import render_to_string
        from django.core.mail import EmailMultiAlternatives
        
        subject = f'Order Approved - {order.order_number}'
        
        # Get site URL
        site_url = request.build_absolute_uri('/').rstrip('/')
        
        # Render HTML email
        html_content = render_to_string('emails/order_approved.html', {
            'order': order,
            'approved_by': request.user.get_full_name() or request.user.username,
            'site_url': site_url,
        })
        
        # Plain text version
        text_content = f'''
        Dear {order.user.get_full_name() or order.user.username},
        
        Good news! Your order {order.order_number} has been approved and is now being processed.
        
        Order Details:
        - Order Number: {order.order_number}
        - Total Amount: ₹{order.total_amount}
        - Status: Processing
        
        Thank you for shopping with us!
        
        Best regards,
        FashioHub Team
        '''
        
        # Send email
        email = EmailMultiAlternatives(subject, text_content, settings.EMAIL_HOST_USER, [order.user.email])
        email.attach_alternative(html_content, "text/html")
        email.send()
        
    except Exception as e:
        print(f'Email sending failed: {e}')
    
    messages.success(request, f'Order {order.order_number} approved successfully!')
    return redirect(request.META.get('HTTP_REFERER', 'admin_orders'))


@login_required(login_url='login')
@staff_member_required(login_url='login')
def admin_reject_order(request, order_id):
    """Reject a pending order"""
    order = get_object_or_404(Order, id=order_id)
    
    if order.approval_status != 'PENDING_APPROVAL':
        messages.warning(request, 'Order is not pending approval.')
        return redirect(request.META.get('HTTP_REFERER', 'admin_orders'))
    
    # Get rejection reason from POST
    if request.method == 'POST':
        rejection_reason = request.POST.get('rejection_reason', '')
        additional_notes = request.POST.get('additional_notes', '')
        
        if not rejection_reason:
            messages.error(request, 'Please select a rejection reason.')
            return redirect(request.META.get('HTTP_REFERER', 'admin_orders'))
        
        # Combine reason and notes
        full_rejection_message = rejection_reason
        if additional_notes:
            full_rejection_message += f"\n\nAdditional Notes: {additional_notes}"
    else:
        full_rejection_message = 'Order rejected by admin'
    
    # Update approval status
    from django.utils import timezone
    order.approval_status = 'REJECTED'
    order.approved_by = request.user
    order.approved_at = timezone.now()
    order.order_status = 'CANCELLED'
    order.approval_notes = f'Rejected by {request.user.username}.\n\nReason: {full_rejection_message}'
    order.save()
    
    # Send rejection email to customer
    try:
        from django.template.loader import render_to_string
        from django.core.mail import EmailMultiAlternatives
        
        subject = f'Order Rejected - {order.order_number}'
        
        # Format reason for email (remove emojis)
        reason_display = rejection_reason
        if '🚫' in rejection_reason or '📦' in rejection_reason or '🚨' in rejection_reason:
            reason_display = rejection_reason.split(' ', 1)[-1] if ' ' in rejection_reason else rejection_reason
        
        # Render HTML email
        html_content = render_to_string('emails/order_rejected.html', {
            'order': order,
            'rejection_reason': reason_display,
            'additional_notes': additional_notes if request.method == 'POST' else '',
        })
        
        # Plain text version
        text_content = f'''
Dear {order.user.get_full_name() or order.user.username},

We regret to inform you that your order {order.order_number} could not be processed and has been cancelled.

Reason: {reason_display}
{f"Additional Information: {additional_notes}" if additional_notes else ""}

Order Details:
- Order Number: {order.order_number}
- Total Amount: ₹{order.total_amount}
- Status: Cancelled

If you have any questions or concerns, please contact our customer support.

We apologize for any inconvenience caused and hope to serve you better in the future.

Thank you for your understanding.

Best regards,
FashioHub Team
        '''
        
        # Send email
        email = EmailMultiAlternatives(subject, text_content, settings.EMAIL_HOST_USER, [order.user.email])
        email.attach_alternative(html_content, "text/html")
        email.send()
        
    except Exception as e:
        print(f'Email sending failed: {e}')
    
    messages.warning(request, f'Order {order.order_number} rejected. Reason: {rejection_reason}')
    return redirect(request.META.get('HTTP_REFERER', 'admin_orders'))


@login_required(login_url='login')
@staff_member_required(login_url='login')
def admin_invoices(request):
    """Admin Invoices List - Using Orders as Invoices"""
    all_invoices = Order.objects.select_related('user').prefetch_related('items').order_by('-created_at')
    
    # Status filter
    status_filter = request.GET.get('status', '')
    search_query = request.GET.get('search', '')
    
    # Filter by status if provided
    if status_filter:
        if status_filter == 'PAID':
            all_invoices = all_invoices.filter(payment_status='PAID')
        elif status_filter == 'PENDING':
            all_invoices = all_invoices.filter(payment_status='PENDING')
        elif status_filter == 'UNPAID':
            all_invoices = all_invoices.filter(payment_status='FAILED')
    
    # Filter by search query
    if search_query:
        all_invoices = all_invoices.filter(
            Q(order_number__icontains=search_query) |
            Q(user__username__icontains=search_query) |
            Q(user__email__icontains=search_query)
        )
    
    # Get statistics
    total_invoices = Order.objects.count()
    total_clients = Order.objects.values('user').distinct().count()
    total_paid = Order.objects.filter(payment_status='PAID').aggregate(Sum('total_amount'))['total_amount__sum'] or 0
    total_unpaid = Order.objects.filter(payment_status__in=['PENDING', 'FAILED']).aggregate(Sum('total_amount'))['total_amount__sum'] or 0
    
    # Pagination
    items_per_page = request.GET.get('per_page', 10)
    try:
        items_per_page = int(items_per_page)
    except (ValueError, TypeError):
        items_per_page = 10
    
    paginator = Paginator(all_invoices, items_per_page)
    page_number = request.GET.get('page', 1)
    
    try:
        page_obj = paginator.page(page_number)
    except PageNotAnInteger:
        page_obj = paginator.page(1)
    except EmptyPage:
        page_obj = paginator.page(paginator.num_pages)
    
    # Build query string for pagination
    query_params = {}
    if search_query:
        query_params['search'] = search_query
    if status_filter:
        query_params['status'] = status_filter
    
    query_string = '&' + urlencode(query_params) if query_params else ''
    
    context = {
        'invoices': page_obj.object_list,
        'page_obj': page_obj,
        'total_invoices': total_invoices,
        'total_clients': total_clients,
        'total_paid': total_paid,
        'total_unpaid': total_unpaid,
        'total_count': all_invoices.count(),
        'status_filter': status_filter,
        'search_query': search_query,
        'items_per_page': items_per_page,
        'query_string': query_string,
    }
    
    return render(request, 'admin_panel/invoices.html', context)


@login_required(login_url='login')
@staff_member_required(login_url='login')
def admin_invoice_inventory(request):
    """Unified Invoice & Inventory management dashboard for admins"""

    invoices_qs = (
        Order.objects
        .select_related('user')
        .prefetch_related('items')
        .order_by('-created_at')
    )

    status_filter = request.GET.get('status', 'all')
    payment_filter = request.GET.get('payment', 'all')
    search_query = request.GET.get('search', '').strip()
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')

    if status_filter and status_filter.lower() != 'all':
        invoices_qs = invoices_qs.filter(order_status=status_filter)

    if payment_filter and payment_filter.lower() != 'all':
        invoices_qs = invoices_qs.filter(payment_status=payment_filter)

    if search_query:
        invoices_qs = invoices_qs.filter(
            Q(order_number__icontains=search_query)
            | Q(user__username__icontains=search_query)
            | Q(user__email__icontains=search_query)
            | Q(user__first_name__icontains=search_query)
            | Q(user__last_name__icontains=search_query)
        )

    if date_from:
        invoices_qs = invoices_qs.filter(created_at__date__gte=date_from)

    if date_to:
        invoices_qs = invoices_qs.filter(created_at__date__lte=date_to)

    items_per_page = request.GET.get('per_page', 10)
    try:
        items_per_page = int(items_per_page)
    except (TypeError, ValueError):
        items_per_page = 10

    paginator = Paginator(invoices_qs, items_per_page)
    page_number = request.GET.get('page', 1)

    try:
        page_obj = paginator.page(page_number)
    except PageNotAnInteger:
        page_obj = paginator.page(1)
    except EmptyPage:
        page_obj = paginator.page(paginator.num_pages)

    paid_orders_qs = Order.objects.filter(payment_status='PAID')
    outstanding_orders_qs = Order.objects.filter(payment_status__in=['PENDING', 'FAILED'])

    total_revenue = paid_orders_qs.aggregate(total=Sum('total_amount'))['total'] or Decimal('0.00')
    outstanding_amount = outstanding_orders_qs.aggregate(total=Sum('total_amount'))['total'] or Decimal('0.00')
    average_order_value = paid_orders_qs.aggregate(avg=Avg('total_amount'))['avg'] or Decimal('0.00')
    overdue_invoices = outstanding_orders_qs.filter(order_status__in=['PENDING', 'PROCESSING']).count()

    low_stock_threshold = 10
    critical_stock_threshold = 3

    inventory_value_expr = ExpressionWrapper(
        F('stock') * F('price'),
        output_field=DecimalField(max_digits=12, decimal_places=2)
    )

    active_products_qs = Product.objects.filter(is_active=True)

    total_skus = active_products_qs.count()
    total_units = active_products_qs.aggregate(total=Sum('stock'))['total'] or 0
    total_inventory_value = active_products_qs.annotate(value=inventory_value_expr).aggregate(total=Sum('value'))['total'] or Decimal('0.00')

    low_stock_qs = (
        active_products_qs
        .filter(stock__gt=0, stock__lte=low_stock_threshold)
        .annotate(value=inventory_value_expr)
        .order_by('stock')
    )
    low_stock_products = list(low_stock_qs[:12])

    out_of_stock_qs = (
        Product.objects
        .filter(stock__lte=0)
        .annotate(value=inventory_value_expr)
        .order_by('name')
    )
    out_of_stock_products = list(out_of_stock_qs[:12])

    low_stock_count = low_stock_qs.count()
    out_of_stock_count = out_of_stock_qs.count()

    top_movers = list(
        active_products_qs
        .annotate(total_sold=Sum('orderitem__quantity'))
        .filter(total_sold__gt=0)
        .order_by('-total_sold')[:8]
    )

    recent_restock = []
    if any(field.name == 'updated_at' for field in Product._meta.get_fields()):
        recent_restock = list(
            active_products_qs
            .filter(updated_at__isnull=False)
            .order_by('-updated_at')[:6]
        )

    # Inventory manager list (all products)
    inventory_qs = Product.objects.all().order_by('name')
    product_search_query = request.GET.get('product_search', '').strip()

    if product_search_query:
        inventory_qs = inventory_qs.filter(
            Q(name__icontains=product_search_query)
            | Q(sku__icontains=product_search_query)
            | Q(brand__icontains=product_search_query)
        )

    product_items_per_page = request.GET.get('product_per_page', 15)
    try:
        product_items_per_page = int(product_items_per_page)
    except (TypeError, ValueError):
        product_items_per_page = 15

    product_paginator = Paginator(inventory_qs, product_items_per_page)
    product_page_number = request.GET.get('product_page', 1)

    try:
        product_page_obj = product_paginator.page(product_page_number)
    except PageNotAnInteger:
        product_page_obj = product_paginator.page(1)
    except EmptyPage:
        product_page_obj = product_paginator.page(product_paginator.num_pages)

    query_params = request.GET.copy()
    query_params.pop('page', None)
    querystring = query_params.urlencode()

    product_query_params = request.GET.copy()
    product_query_params.pop('product_page', None)
    product_querystring = product_query_params.urlencode()

    context = {
        'page_obj': page_obj,
        'invoices': page_obj.object_list,
        'items_per_page': items_per_page,
        'status_filter': status_filter,
        'payment_filter': payment_filter,
        'search_query': search_query,
        'date_from': date_from,
        'date_to': date_to,
        'querystring': querystring,
        'product_querystring': product_querystring,
        'current_path': request.get_full_path(),

        'invoice_total_count': invoices_qs.count(),
        'paid_invoice_count': paid_orders_qs.count(),
        'outstanding_invoice_count': outstanding_orders_qs.count(),
        'total_revenue': total_revenue,
        'outstanding_amount': outstanding_amount,
        'average_order_value': average_order_value,
        'overdue_invoices': overdue_invoices,

        'low_stock_products': low_stock_products,
        'out_of_stock_products': out_of_stock_products,
        'low_stock_threshold': low_stock_threshold,
        'critical_stock_threshold': critical_stock_threshold,
        'low_stock_count': low_stock_count,
        'out_of_stock_count': out_of_stock_count,
        'total_skus': total_skus,
        'total_units': total_units,
        'total_inventory_value': total_inventory_value,
        'top_movers': top_movers,
        'recent_restock': recent_restock,
        'status_choices': Order.ORDER_STATUS_CHOICES,
        'payment_status_choices': Order.PAYMENT_STATUS_CHOICES,
        'product_page_obj': product_page_obj,
        'inventory_products': product_page_obj.object_list,
        'product_items_per_page': product_items_per_page,
        'product_search_query': product_search_query,
    }

    return render(request, 'admin_panel/invoice_inventory.html', context)


@login_required(login_url='login')
@staff_member_required(login_url='login')
@require_POST
def admin_update_inventory(request):
    """Adjust product stock from the Inventory dashboard."""
    product_id = request.POST.get('product_id')
    action = request.POST.get('action', 'save_stock')
    redirect_url = request.POST.get('next') or reverse('admin_invoice_inventory')

    try:
        product = Product.objects.get(id=product_id)
    except Product.DoesNotExist:
        messages.error(request, 'Product not found')
        return redirect(redirect_url)

    try:
        previous_stock = product.stock
        if action == 'mark_out':
            product.stock = 0
            # Keep the product visible with an out-of-stock tag
            product.is_active = True
            product.save(update_fields=['stock', 'is_active'])
            messages.success(request, f'Marked {product.name} as out of stock (visible in shop).')
        else:
            raw_stock = request.POST.get('stock', '0')
            new_stock = max(0, int(raw_stock))
            product.stock = new_stock

            # Reactivate when restocked; keep visible even at 0
            product.is_active = True if new_stock >= 0 else product.is_active
            product.save(update_fields=['stock', 'is_active'])
            messages.success(request, f'Updated stock for {product.name} to {new_stock}.')

        # Notify subscribers when restocked from zero
        if previous_stock <= 0 and product.stock > 0:
            notifications = ProductStockNotification.objects.filter(product=product, is_sent=False)
            sent_count = 0
            failed_count = 0
            last_error = ''
            for note in notifications:
                try:
                    send_mail(
                        subject=f"{product.name} is back in stock",
                        message=f"Good news! {product.name} is available again. Visit the product page to purchase.",
                        from_email=getattr(settings, 'DEFAULT_FROM_EMAIL', None),
                        recipient_list=[note.email],
                        fail_silently=False,
                    )
                    note.mark_sent()
                    sent_count += 1
                except Exception as exc:
                    failed_count += 1
                    last_error = str(exc)
            if notifications.exists():
                messages.info(request, f"Restock notifications attempted: sent {sent_count}, failed {failed_count}.")
                if failed_count and last_error:
                    messages.warning(request, f"Email error: {last_error}")
            else:
                messages.info(request, "No pending restock notifications for this product.")
    except ValueError:
        messages.error(request, 'Invalid stock value')
    except Exception as exc:
        messages.error(request, f'Could not update stock: {exc}')

    return redirect(redirect_url)


@login_required
def admin_customers(request):
    """Admin view to manage customers"""
    if not request.user.is_staff:
        messages.error(request, 'Unauthorized access')
        return redirect('index')
    
    # Handle POST requests (block/unblock, segment update)
    if request.method == 'POST':
        action = request.POST.get('action')
        customer_ids = request.POST.getlist('customer_ids')
        
        if action and customer_ids:
            users = User.objects.filter(id__in=customer_ids)
            
            if action == 'block':
                UserProfile.objects.filter(user__in=users).update(is_blocked=True)
                messages.success(request, f'Successfully blocked {len(customer_ids)} customer(s)')
            elif action == 'unblock':
                UserProfile.objects.filter(user__in=users).update(is_blocked=False)
                messages.success(request, f'Successfully unblocked {len(customer_ids)} customer(s)')
            elif action in ['NEW', 'REGULAR', 'VIP', 'ADMIN']:
                UserProfile.objects.filter(user__in=users).update(customer_segment=action)
                users.update(is_staff=(action == 'ADMIN'))
                messages.success(request, f'Successfully updated {len(customer_ids)} customer(s) to {action}')
            elif action == 'delete':
                users.delete()
                messages.success(request, f'Successfully deleted {len(customer_ids)} customer(s)')
            
            return redirect('admin_customers')
    
    # Get filter parameters
    segment_filter = request.GET.get('segment', 'all')
    status_filter = request.GET.get('status', 'all')
    search_query = request.GET.get('search', '')
    
    # Base queryset - include all users
    customers = User.objects.all().select_related('userprofile')
    
    # Apply filters
    if segment_filter != 'all':
        customers = customers.filter(userprofile__customer_segment=segment_filter)
    
    if status_filter == 'blocked':
        customers = customers.filter(userprofile__is_blocked=True)
    elif status_filter == 'active':
        customers = customers.filter(userprofile__is_blocked=False)
    
    if search_query:
        customers = customers.filter(
            Q(username__icontains=search_query) |
            Q(email__icontains=search_query) |
            Q(first_name__icontains=search_query) |
            Q(last_name__icontains=search_query)
        )

    # Always show ADMIN segment users first
    customers = customers.annotate(
        admin_order=Case(
            When(userprofile__customer_segment='ADMIN', then=Value(0)),
            default=Value(1),
            output_field=IntegerField(),
        )
    ).order_by('admin_order', 'id')
    
    # Calculate stats
    total_customers = User.objects.all().count()  # Include all users
    active_customers = UserProfile.objects.filter(is_blocked=False).count()
    blocked_customers = UserProfile.objects.filter(is_blocked=True).count()
    new_customers = UserProfile.objects.filter(customer_segment='NEW').count()
    regular_customers = UserProfile.objects.filter(customer_segment='REGULAR').count()
    vip_customers = UserProfile.objects.filter(customer_segment='VIP').count()
    admin_customers = UserProfile.objects.filter(customer_segment='ADMIN').count()
    
    # Pagination
    paginator = Paginator(customers, 20)
    page_number = request.GET.get('page', 1)
    
    try:
        customers_page = paginator.page(page_number)
    except PageNotAnInteger:
        customers_page = paginator.page(1)
    except EmptyPage:
        customers_page = paginator.page(paginator.num_pages)
    
    context = {
        'customers': customers_page,
        'total_customers': total_customers,
        'active_customers': active_customers,
        'blocked_customers': blocked_customers,
        'new_customers': new_customers,
        'regular_customers': regular_customers,
        'vip_customers': vip_customers,
        'admin_customers': admin_customers,
        'segment_filter': segment_filter,
        'status_filter': status_filter,
        'search_query': search_query,
    }
    return render(request, 'admin_panel/customers.html', context)


@login_required
def admin_customer_details(request, customer_id):
    """Admin view to see comprehensive customer details with real-time data"""
    if not request.user.is_staff:
        messages.error(request, 'Unauthorized access')
        return redirect('index')
    
    customer = get_object_or_404(User, id=customer_id)
    profile, created = UserProfile.objects.get_or_create(user=customer)
    
    # Handle POST requests (update segment, block/unblock, admin role)
    if request.method == 'POST':
        action = request.POST.get('action')
        
        if action == 'toggle_block':
            profile.is_blocked = not profile.is_blocked
            profile.save()
            status = 'blocked' if profile.is_blocked else 'unblocked'
            messages.success(request, f'Customer {status} successfully')
        elif action == 'update_segment':
            new_segment = request.POST.get('segment')
            if new_segment in ['NEW', 'REGULAR', 'VIP', 'ADMIN']:
                profile.customer_segment = new_segment
                customer.is_staff = (new_segment == 'ADMIN')
                customer.save()
                profile.save()
                messages.success(request, f'Customer segment updated to {new_segment}')
        elif action == 'toggle_admin':
            customer.is_staff = not customer.is_staff
            customer.save()
            if customer.is_staff:
                profile.customer_segment = 'ADMIN'
            else:
                profile.customer_segment = 'REGULAR'
            profile.save()
            status = 'granted' if customer.is_staff else 'revoked'
            messages.success(request, f'Admin access {status} successfully')
        
        return redirect('admin_customer_details', customer_id=customer_id)
    
    # Wishlist data
    wishlist_items = Wishlist.objects.filter(user=customer).select_related('product')
    total_wishlist = wishlist_items.count()
    
    # Cart data
    cart_items = Cart.objects.filter(user=customer).select_related('product')
    total_cart_value = sum(item.get_total_price() for item in cart_items)
    
    # Reviews data
    reviews = ProductReview.objects.filter(user=customer).select_related('product').order_by('-created_at')
    total_reviews = reviews.count()
    approved_reviews = reviews.filter(is_approved=True).count()
    
    # Orders data
    orders = Order.objects.filter(user=customer).order_by('-created_at')
    total_orders = orders.count()
    
    # Calculate total spent (from PAID orders)
    paid_orders = orders.filter(payment_status='PAID')
    total_spent = paid_orders.aggregate(total=Sum('total_amount'))['total'] or Decimal('0.00')
    
    # Update profile total_spent
    if profile.total_spent != total_spent:
        profile.total_spent = total_spent
        profile.save(update_fields=['total_spent'])
    
    # Calculate profit (assuming 30% profit margin on paid orders)
    total_profit = total_spent * Decimal('0.30')
    
    # Order statistics
    pending_orders = orders.filter(order_status='PENDING').count()
    processing_orders = orders.filter(order_status='PROCESSING').count()
    shipped_orders = orders.filter(order_status='SHIPPED').count()
    delivered_orders = orders.filter(order_status='DELIVERED').count()
    cancelled_orders = orders.filter(order_status='CANCELLED').count()
    
    # Payment statistics
    paid_orders_count = orders.filter(payment_status='PAID').count()
    pending_payments = orders.filter(payment_status='PENDING').count()
    
    # Products purchased
    order_items = OrderItem.objects.filter(order__user=customer).select_related('product', 'order')
    unique_products = order_items.values('product').distinct().count()
    total_items_purchased = order_items.aggregate(total=Sum('quantity'))['total'] or 0
    
    # Top purchased products
    from collections import Counter
    product_purchases = Counter()
    for item in order_items:
        if item.product:
            product_purchases[item.product] += item.quantity
    top_products = [{'product': product, 'count': count} for product, count in product_purchases.most_common(5)]
    
    # Recent activity logs - Build comprehensive activity timeline
    from datetime import timedelta
    recent_activities = []
    
    # Add all individual wishlist additions
    for wishlist_item in wishlist_items.order_by('-added_at')[:10]:
        recent_activities.append({
            'icon': 'bx bx-heart',
            'color': 'danger',
            'title': 'Added to Wishlist',
            'description': f'{wishlist_item.product.name}',
            'timestamp': wishlist_item.added_at
        })
    
    # Add all individual cart additions
    for cart_item in cart_items.order_by('-added_at')[:10]:
        recent_activities.append({
            'icon': 'bx bx-cart-add',
            'color': 'info',
            'title': 'Added to Cart',
            'description': f'{cart_item.product.name} (Qty: {cart_item.quantity})',
            'timestamp': cart_item.added_at
        })
    
    # Add all individual reviews
    for review in reviews[:10]:
        recent_activities.append({
            'icon': 'bx bx-message-square-detail',
            'color': 'warning',
            'title': 'Submitted Review',
            'description': f'{review.product.name} - {review.rating} stars {"✓ Approved" if review.is_approved else "⏳ Pending"}',
            'timestamp': review.created_at
        })
    
    # Add all orders with detailed status
    for order in orders[:15]:
        status_icons = {
            'PENDING': 'bx bx-time-five',
            'PROCESSING': 'bx bx-loader-alt',
            'SHIPPED': 'bx bx-package',
            'DELIVERED': 'bx bx-check-circle',
            'CANCELLED': 'bx bx-x-circle',
        }
        status_colors = {
            'PENDING': 'warning',
            'PROCESSING': 'info',
            'SHIPPED': 'primary',
            'DELIVERED': 'success',
            'CANCELLED': 'danger',
        }
        recent_activities.append({
            'icon': status_icons.get(order.order_status, 'bx bx-shopping-bag'),
            'color': status_colors.get(order.order_status, 'primary'),
            'title': f'Order #{order.order_number}',
            'description': f'₹{order.total_amount} - {order.get_order_status_display()} ({order.get_payment_status_display()})',
            'timestamp': order.created_at
        })
    
    # Account created
    recent_activities.append({
        'icon': 'bx bx-user-plus',
        'color': 'success',
        'title': 'Account Created',
        'description': 'Joined VibeMall',
        'timestamp': customer.date_joined
    })
    
    # Last login/activity
    if profile.last_activity and profile.last_activity > customer.date_joined:
        recent_activities.append({
            'icon': 'bx bx-log-in',
            'color': 'info',
            'title': 'Last Activity',
            'description': 'User active on site',
            'timestamp': profile.last_activity
        })
    
    # Sort activities by timestamp (newest first)
    recent_activities.sort(key=lambda x: x['timestamp'], reverse=True)
    recent_activities = recent_activities[:25]  # Limit to 25 most recent
    
    context = {
        'customer': customer,
        'profile': profile,
        
        # Stats cards
        'total_spent': total_spent,
        'total_wishlist': total_wishlist,
        'total_cart_value': total_cart_value,
        'total_reviews': total_reviews,
        'approved_reviews': approved_reviews,
        
        # Tab content
        'wishlist_items': wishlist_items,
        'cart_items': cart_items,
        'reviews': reviews,
        'recent_activities': recent_activities,
        
        # Orders & Profit section
        'orders': orders[:10],  # Latest 10 orders
        'total_orders': total_orders,
        'total_profit': total_profit,
        'pending_orders': pending_orders,
        'processing_orders': processing_orders,
        'shipped_orders': shipped_orders,
        'delivered_orders': delivered_orders,
        'cancelled_orders': cancelled_orders,
        'paid_orders_count': paid_orders_count,
        'pending_payments': pending_payments,
        
        # Products purchased
        'unique_products': unique_products,
        'total_items_purchased': total_items_purchased,
        'top_products': top_products,
        'order_items': order_items[:10],  # Latest 10 items
        
        # Segment choices for dropdown
        'segment_choices': UserProfile.CUSTOMER_SEGMENT_CHOICES,
    }
    return render(request, 'admin_panel/customer_details.html', context)


# ===== PUBLIC VIEWS =====

def index(request):
    sliders = Slider.objects.filter(is_active=True)
    features = Feature.objects.filter(is_active=True)
    banners = Banner.objects.filter(is_active=True)
    categories = CategoryIcon.objects.filter(is_active=True)
    
    # Get latest active products for all sections
    # Order by -id since newer products have higher IDs
    latest_products = Product.objects.filter(is_active=True).order_by('-id')
    top_deals = latest_products[:10]
    top_selling = latest_products[:10]
    top_featured = latest_products[:10]
    recommended = latest_products[:10]
    
    countdown = DealCountdown.objects.filter(is_active=True).first()
    brand_partners = BrandPartner.objects.filter(is_active=True).order_by('order')

    # Get wishlist product IDs for logged-in user
    wishlist_product_ids = []
    cart_product_ids = []
    delivered_orders = []
    
    if request.user.is_authenticated:
        wishlist_product_ids = list(Wishlist.objects.filter(user=request.user).values_list('product_id', flat=True))
        cart_product_ids = list(Cart.objects.filter(user=request.user).values_list('product_id', flat=True))
        
        # Check for recently delivered orders (delivered in last 7 days without reviews)
        from datetime import timedelta
        from django.utils import timezone
        seven_days_ago = timezone.now() - timedelta(days=7)
        
        # Get delivered orders that don't have review notification shown
        shown_review_notifications = request.session.get('shown_review_notifications', [])
        
        delivered_orders = Order.objects.filter(
            user=request.user,
            order_status='DELIVERED',
            delivery_date__gte=seven_days_ago
        ).exclude(
            order_number__in=shown_review_notifications
        ).prefetch_related('items__product')[:3]  # Show up to 3 recent delivered orders
        
        # Mark these orders as shown
        if delivered_orders:
            for order in delivered_orders:
                shown_review_notifications.append(order.order_number)
            request.session['shown_review_notifications'] = shown_review_notifications
            request.session.modified = True

    return render(
        request,
        'index.html',
        {
            'sliders': sliders,
            'features': features,
            'banners': banners,
            'categories': categories,
            'top_deals': top_deals,
            'top_selling': top_selling,
            'top_featured': top_featured,
            'recommended': recommended,
            'countdown': countdown,
            'wishlist_product_ids': wishlist_product_ids,
            'cart_product_ids': cart_product_ids,
            'brand_partners': brand_partners,
            'delivered_orders': delivered_orders,
        }
    )










def about(request): return render(request, 'about.html')
def blog(request): return render(request, 'blog.html')
def blog_details(request): return render(request, 'blog-details.html')


@login_required(login_url='login')
def cart(request):
    cart_items = Cart.objects.filter(user=request.user).select_related('product')
    total_price = sum(item.get_total_price() for item in cart_items)
    cart_count = cart_items.count()
    
    return render(request, 'cart.html', {
        'cart_items': cart_items,
        'total_price': total_price,
        'cart_count': cart_count
    })



def buy_now(request, product_id):
    """Buy Now - Redirect directly to checkout with product"""
    try:
        product = Product.objects.get(id=product_id, is_active=True)
        quantity = int(request.POST.get('quantity', 1))
        
        if quantity < 1:
            return JsonResponse({
                'success': False,
                'message': 'Invalid quantity'
            }, status=400)
        
        # Check stock
        if product.stock < quantity:
            return JsonResponse({
                'success': False,
                'message': f'Only {product.stock} items available in stock'
            }, status=400)
        
        # Store in session for checkout
        request.session['buy_now_item'] = {
            'product_id': product.id,
            'quantity': quantity,
            'price': str(product.price)
        }
        
        return JsonResponse({
            'success': True,
            'message': f'{product.name} added to checkout!',
            'redirect_url': '/checkout/'
        })
        
    except Product.DoesNotExist:
        return JsonResponse({
            'success': False,
            'message': 'Product not found'
        }, status=404)


@login_required(login_url='login')
def checkout(request):
    """Dynamic Checkout Page"""
    # Get cart items OR buy_now item
    cart_items = []
    buy_now_item = None
    total_price = 0
    
    # Check if there's a buy_now item in session
    if 'buy_now_item' in request.session:
        buy_now_data = request.session['buy_now_item']
        try:
            product = Product.objects.get(id=buy_now_data['product_id'])
            buy_now_item = {
                'product': product,
                'quantity': buy_now_data['quantity'],
                'price': float(buy_now_data['price']),
                'subtotal': float(buy_now_data['price']) * buy_now_data['quantity']
            }
            total_price = buy_now_item['subtotal']
            cart_items = [buy_now_item]
        except Product.DoesNotExist:
            del request.session['buy_now_item']
    else:
        # Get items from cart
        cart_items = Cart.objects.filter(user=request.user).select_related('product')
        total_price = sum(item.get_total_price() for item in cart_items)
    
    # Handle form submission
    if request.method == 'POST':
        # Get form data
        first_name = request.POST.get('first_name')
        last_name = request.POST.get('last_name')
        email = request.POST.get('email')
        phone = request.POST.get('phone')
        country = request.POST.get('country')
        address = request.POST.get('address')
        city = request.POST.get('city')
        state = request.POST.get('state')
        postcode = request.POST.get('postcode')
        payment_method = request.POST.get('payment_method', 'COD')
        customer_notes = request.POST.get('customer_notes', '')
        is_resell = request.POST.get('is_resell') == 'on'
        
        # Loyalty points redemption
        redeem_points = request.POST.get('redeem_points') == 'on'
        points_to_redeem = 0
        if redeem_points:
            points_to_redeem = int(request.POST.get('points_to_redeem', 0))
        
        # Debug: Show selected payment method
        print(f"[DEBUG] Selected payment method: {payment_method}")
        if not payment_method or payment_method not in ['COD', 'RAZORPAY']:
            messages.error(request, 'Please select a valid payment method.')
            return render(request, 'checkout.html', {
                'cart_items': cart_items,
                'total_price': total_price,
                'buy_now_item': buy_now_item
            })
        
        # Resell FROM details
        from_name = request.POST.get('from_name', '')
        from_phone = request.POST.get('from_phone', '')
        
        # Validation
        if not all([first_name, last_name, email, phone, address, city, state, postcode]):
            messages.error(request, 'Please fill all required fields')
            return render(request, 'checkout.html', {
                'cart_items': cart_items,
                'total_price': total_price,
                'buy_now_item': buy_now_item
            })
        
        # Validate resell FROM details if resell order
        if is_resell and (not from_name or not from_phone):
            messages.error(request, 'Please provide FROM details for resell order')
            return render(request, 'checkout.html', {
                'cart_items': cart_items,
                'total_price': total_price,
                'buy_now_item': buy_now_item
            })
        
        # Create order
        try:
            # Calculate totals
            subtotal = Decimal(str(total_price))
            tax = subtotal * Decimal('0.05')  # 5% tax
            shipping_cost = Decimal('0.00') if subtotal > 500 else Decimal('50.00')  # Free shipping above 500
            
            # Apply loyalty points discount (1 point = ₹0.03)
            points_discount = Decimal('0')
            if redeem_points and points_to_redeem > 0:
                # Verify user has enough points
                loyalty_account = LoyaltyPoints.objects.get(user=request.user)
                if points_to_redeem <= loyalty_account.points_available:
                    points_discount = Decimal(str(points_to_redeem)) * Decimal('0.03')
                else:
                    messages.error(request, 'Insufficient loyalty points')
                    return render(request, 'checkout.html', {
                        'cart_items': cart_items,
                        'total_price': total_price,
                        'buy_now_item': buy_now_item
                    })
            
            total_amount = subtotal + tax + shipping_cost - points_discount
            
            # Ensure total doesn't go negative
            if total_amount < 0:
                total_amount = Decimal('0')
            
            # Create shipping address
            shipping_address = f"{first_name} {last_name}\n{address}\n{city}, {state} {postcode}\n{country}"
            
            # Create order
            order = Order.objects.create(
                user=request.user,
                subtotal=subtotal,
                tax=tax,
                shipping_cost=shipping_cost,
                total_amount=total_amount,
                shipping_address=shipping_address,
                billing_address=shipping_address,
                payment_method=payment_method,
                customer_notes=customer_notes,
                is_resell=is_resell,
                resell_from_name=from_name if is_resell else '',
                resell_from_phone=from_phone if is_resell else ''
            )
            
            # Store points redemption info for later
            if redeem_points and points_to_redeem > 0:
                order.admin_notes += f"\nLoyalty Points Redeemed: {points_to_redeem} points (₹{points_discount} discount)"
                order.save()
            
            # Add order items
            if buy_now_item:
                # From buy_now
                product = buy_now_item['product']
                # Build absolute image URL
                product_image = ''
                if product.image:
                    image_url = product.image.url
                    if not image_url.startswith('http'):
                        site_url = request.build_absolute_uri('/').rstrip('/')
                        product_image = f"{site_url}{image_url}"
                    else:
                        product_image = image_url
                
                OrderItem.objects.create(
                    order=order,
                    product=product,
                    product_name=product.name,
                    product_price=buy_now_item['price'],
                    product_image=product_image,
                    quantity=buy_now_item['quantity']
                )
                # Clear buy_now from session
                del request.session['buy_now_item']
            else:
                # From cart
                for item in cart_items:
                    # Build absolute image URL
                    product_image = ''
                    if item.product.image:
                        image_url = item.product.image.url
                        if not image_url.startswith('http'):
                            site_url = request.build_absolute_uri('/').rstrip('/')
                            product_image = f"{site_url}{image_url}"
                        else:
                            product_image = image_url
                    
                    OrderItem.objects.create(
                        order=order,
                        product=item.product,
                        product_name=item.product.name,
                        product_price=item.product.price,
                        product_image=product_image,
                        quantity=item.quantity
                    )
            
            # Redeem loyalty points if applicable
            if redeem_points and points_to_redeem > 0:
                loyalty_account = LoyaltyPoints.objects.get(user=request.user)
                loyalty_account.redeem_points(points_to_redeem, f"Order #{order.order_number} - ₹{points_discount} discount")
            
            # Handle payment
            if payment_method == 'RAZORPAY':
                # Redirect to payment gateway
                return redirect('razorpay_payment', order_id=order.id)
            else:
                # COD - Mark as pending
                order.payment_status = 'PENDING'
                order.order_status = 'PENDING'
                order.save()
                
                # Auto-process approval (fraud detection & auto-approve)
                order.auto_process_approval()
                
                # Clear cart after successful order
                if not buy_now_item:
                    Cart.objects.filter(user=request.user).delete()
                
                # Send order confirmation email to customer
                send_order_confirmation_email(order)
                
                # Send notification email to admin
                send_admin_order_notification(order, request)
                
                # Check if order needs approval
                if order.approval_status == 'PENDING_APPROVAL':
                    messages.warning(request, f'Order placed successfully! Order #: {order.order_number}. Your order is pending approval due to security checks.')
                else:
                    messages.success(request, f'Order placed successfully! Order #: {order.order_number}')
                
                return redirect('order_confirmation', order_id=order.id)
                
        except Exception as e:
            messages.error(request, f'Error creating order: {str(e)}')
            return render(request, 'checkout.html', {
                'cart_items': cart_items,
                'total_price': total_price,
                'buy_now_item': buy_now_item
            })
    
    # GET request - Show checkout form
    user_profile = UserProfile.objects.filter(user=request.user).first()
    
    # Get loyalty account
    loyalty_account = None
    if request.user.is_authenticated:
        try:
            loyalty_account = LoyaltyPoints.objects.get(user=request.user)
        except LoyaltyPoints.DoesNotExist:
            pass
    
    context = {
        'cart_items': cart_items,
        'total_price': total_price,
        'buy_now_item': buy_now_item,
        'user_profile': user_profile,
        'loyalty_account': loyalty_account,
        'tax_amount': Decimal(str(total_price)) * Decimal('0.05'),
        'shipping_cost': Decimal('0.00') if Decimal(str(total_price)) > 500 else Decimal('50.00'),
        'final_total': Decimal(str(total_price)) + (Decimal(str(total_price)) * Decimal('0.05')) + (Decimal('0.00') if Decimal(str(total_price)) > 500 else Decimal('50.00'))
    }
    
    return render(request, 'checkout.html', context)
def contact(request): return render(request, 'contact.html')
def faq(request): return render(request, 'faq.html')


def login_view(request):
    if request.method == "POST":
        username = request.POST.get("username")
        password = request.POST.get("password")
        remember_me = request.POST.get("remember_me")

        user = authenticate(request, username=username, password=password)
        if user:
            login(request, user)
            # Set session expiry based on remember me checkbox
            if remember_me:
                request.session.set_expiry(30 * 24 * 60 * 60)  # 30 days
            else:
                request.session.set_expiry(0)  # Session expires when browser closes
            return redirect("index")
        else:
            messages.error(request, "Invalid username or password. Please check your credentials and try again.")

    return render(request, "login.html")




def my_account(request): return render(request, 'profile.html')
def product(request): return render(request, 'product.html')
def product_details(request, product_id=None):
    """Display product details with dynamic data and enhanced reviews"""
    if product_id:
        try:
            product = Product.objects.get(id=product_id, is_active=True)
            in_wishlist = False
            if request.user.is_authenticated:
                in_wishlist = Wishlist.objects.filter(user=request.user, product=product).exists()

            # Auto-generate reviews for existing products if missing
            if product.review_count and product.rating:
                existing_qs = ProductReview.objects.filter(product=product, is_approved=True)
                existing_count = existing_qs.count()
                if existing_count < product.review_count:
                    existing_sum = existing_qs.aggregate(total=Sum('rating'))['total'] or 0
                    desired_sum = int(round(float(product.rating) * int(product.review_count)))
                    missing_n = int(product.review_count) - existing_count
                    if missing_n > 0:
                        missing_sum = desired_sum - existing_sum
                        missing_avg = missing_sum / missing_n if missing_n else float(product.rating)
                        if missing_avg < 1:
                            missing_avg = 1
                        if missing_avg > 5:
                            missing_avg = 5
                        generate_auto_reviews(product, missing_n, missing_avg, request.user if request.user.is_authenticated else None)
            
            # Get filter and sort parameters
            rating_filter = request.GET.get('rating', 'all')
            sort_by = request.GET.get('sort', 'recent')
            
            # Get approved customer reviews for this product (exclude auto-generated)
            approved_reviews = ProductReview.objects.filter(
                product=product, 
                is_approved=True,
                is_auto_generated=False
            ).select_related('user').prefetch_related('images', 'votes')
            
            # Apply rating filter
            if rating_filter != 'all':
                try:
                    rating_value = int(rating_filter)
                    approved_reviews = approved_reviews.filter(rating=rating_value)
                except ValueError:
                    pass
            
            # Apply sorting
            if sort_by == 'recent':
                approved_reviews = approved_reviews.order_by('-created_at')
            elif sort_by == 'rating_high':
                approved_reviews = approved_reviews.order_by('-rating', '-created_at')
            elif sort_by == 'rating_low':
                approved_reviews = approved_reviews.order_by('rating', '-created_at')
            elif sort_by == 'helpful':
                approved_reviews = approved_reviews.order_by('-helpful_count', '-created_at')
            
            review_count = approved_reviews.count()
            
            # Calculate rating breakdown (include ALL approved reviews for stats)
            total_reviews = ProductReview.objects.filter(product=product, is_approved=True).count()
            rating_breakdown = {
                5: ProductReview.objects.filter(product=product, is_approved=True, rating=5).count(),
                4: ProductReview.objects.filter(product=product, is_approved=True, rating=4).count(),
                3: ProductReview.objects.filter(product=product, is_approved=True, rating=3).count(),
                2: ProductReview.objects.filter(product=product, is_approved=True, rating=2).count(),
                1: ProductReview.objects.filter(product=product, is_approved=True, rating=1).count(),
            }
            
            # Calculate percentages
            rating_percentages = {}
            for rating, count in rating_breakdown.items():
                rating_percentages[rating] = int((count / total_reviews * 100)) if total_reviews > 0 else 0
            
            # Calculate average rating (include ALL approved reviews for stats)
            avg_rating = ProductReview.objects.filter(
                product=product, 
                is_approved=True
            ).aggregate(Avg('rating'))['rating__avg'] or 0
            
            # Get product questions
            approved_questions = ProductQuestion.objects.filter(
                product=product,
                is_approved=True
            ).select_related('user', 'answered_by').order_by('-answered_at', '-created_at')
            
            total_questions = approved_questions.count()
            
            # Get user's votes if authenticated
            user_votes = {}
            if request.user.is_authenticated:
                votes = ReviewVote.objects.filter(
                    user=request.user,
                    review__in=approved_reviews
                ).values_list('review_id', 'is_helpful')
                user_votes = dict(votes)
            
            return render(request, 'product-details.html', {
                'product': product,
                'in_wishlist': in_wishlist,
                'reviews': approved_reviews,
                'review_count': review_count,
                'total_reviews': total_reviews,
                'avg_rating': round(avg_rating, 1),
                'rating_breakdown': rating_breakdown,
                'rating_percentages': rating_percentages,
                'approved_questions': approved_questions,
                'total_questions': total_questions,
                'user_votes': user_votes,
                'rating_filter': rating_filter,
                'sort_by': sort_by,
            })
        except Product.DoesNotExist:
            return render(request, '404.html', status=404)
    else:
        return render(request, '404.html', status=404)
def shop(request):
    products = Product.objects.filter(is_active=True)
    
    # Get banners for shop page
    banners = Banner.objects.filter(is_active=True).filter(
        Q(page_type='SHOP') | Q(page_type='BOTH')
    ).order_by('order')

    # Filters from query params
    min_price = request.GET.get('min_price')
    max_price = request.GET.get('max_price')
    min_rating = request.GET.get('min_rating')

    if min_price:
        try:
            products = products.filter(price__gte=float(min_price))
        except ValueError:
            pass
    if max_price:
        try:
            products = products.filter(price__lte=float(max_price))
        except ValueError:
            pass
    if min_rating:
        try:
            products = products.filter(rating__gte=float(min_rating))
        except ValueError:
            pass

    special_offers = Product.objects.filter(is_active=True).order_by('-discount_percent', '-id')[:5]

    # Get categories from CategoryIcon (dynamic admin-managed categories)
    category_icons = CategoryIcon.objects.filter(is_active=True).order_by('order', 'id')
    # Category data without product counts
    category_data = [
        (cat.category_key, cat.name, 0)
        for cat in category_icons
    ]

    wishlist_product_ids = set()
    cart_product_ids = set()
    if request.user.is_authenticated:
        wishlist_product_ids = set(
            Wishlist.objects.filter(user=request.user).values_list('product_id', flat=True)
        )
        cart_product_ids = set(
            Cart.objects.filter(user=request.user).values_list('product_id', flat=True)
        )

    # Pagination - 16 products per page (4 columns x 4 rows)
    paginator = Paginator(products, 16)
    page = request.GET.get('page', 1)
    
    try:
        products_page = paginator.page(page)
    except PageNotAnInteger:
        products_page = paginator.page(1)
    except EmptyPage:
        products_page = paginator.page(paginator.num_pages)

    return render(request, 'shop.html', {
        'products': products_page,
        'banners': banners,
        'special_offers': special_offers,
        'category_data': category_data,
        'selected_category': '',
        'min_price': min_price or '',
        'max_price': max_price or '',
        'selected_rating': min_rating or '',
        'wishlist_product_ids': wishlist_product_ids,
        'cart_product_ids': cart_product_ids,
    })
def shop_details(request): return render(request, 'shop-details.html')


@login_required(login_url='login')
def wishlist(request):
    wishlist_items = Wishlist.objects.filter(user=request.user).select_related('product')
    wishlist_count = wishlist_items.count()
    
    return render(request, 'wishlist.html', {
        'wishlist_items': wishlist_items,
        'wishlist_count': wishlist_count
    })

def page_404(request): return render(request, '404.html', status=404)
def order_tracking(request):return render(request, 'order-tracking.html')



def register_view(request):
    if request.method == "POST":
        name = request.POST.get("name")
        username = request.POST.get("username")
        email = request.POST.get("email")
        password = request.POST.get("password")
        confirm_password = request.POST.get("confirm_password")
        country_code = request.POST.get("country_code")
        mobile_number = request.POST.get("mobile_number")
        profile_image = request.FILES.get("profile_image")

        # Validate password strength
        if len(password) < 8:
            messages.error(request, "Password must be at least 8 characters long")
            return redirect("register")

        if password != confirm_password:
            messages.error(request, "Passwords do not match. Please ensure both passwords are identical.")
            return redirect("register")

        if User.objects.filter(username=username).exists():
            messages.error(request, "Username already taken. Please choose a different username.")
            return redirect("register")

        if User.objects.filter(email=email).exists():
            messages.error(request, "Email already registered. Please use a different email or login.")
            return redirect("register")

        try:
            user = User.objects.create_user(
                username=username,
                email=email,
                password=password
            )
            user.first_name = name
            user.is_active = False
            user.save()

            profile = user.userprofile
            profile.country_code = country_code
            profile.mobile_number = mobile_number
            if profile_image:
                profile.profile_image = profile_image
            profile.save()

            send_verification_email(user, request)
            messages.success(request, "Account created! Please verify your email to activate your account.", extra_tags='success')
            return redirect("verify_email_sent")
        except Exception as e:
            messages.error(request, f"An error occurred during registration: {str(e)}")
            return redirect("register")

    return render(request, "register.html")


def send_verification_email(user, request):
    """Send email verification link"""
    uid = urlsafe_base64_encode(force_bytes(user.pk))
    token = default_token_generator.make_token(user)
    verify_url = request.build_absolute_uri(
        f"/verify-email/{uid}/{token}/"
    )

    subject = "Verify your FashionHub account"
    message = render_to_string('emails/verify_email.html', {
        'user': user,
        'verify_url': verify_url,
    })

    send_mail(
        subject,
        message,
        settings.DEFAULT_FROM_EMAIL,
        [user.email],
        fail_silently=False,
        html_message=message,
    )


def verify_email_sent(request):
    """Show verification sent page"""
    return render(request, 'verify_email_sent.html')


def verify_email(request, uidb64, token):
    """Verify user email and activate account"""
    try:
        uid = urlsafe_base64_decode(uidb64).decode()
        user = User.objects.get(pk=uid)

        if default_token_generator.check_token(user, token):
            user.is_active = True
            user.save()
            messages.success(request, "Email verified successfully! You can log in now.")
            return redirect('login')

        messages.error(request, "Verification link is invalid or expired.")
        return redirect('register')
    except (TypeError, ValueError, OverflowError, User.DoesNotExist):
        messages.error(request, "Invalid verification link.")
        return redirect('register')

    return render(request, "register.html")


def logout_view(request):
    logout(request)
    return redirect("index")


# ===== PASSWORD RESET VIEWS =====

def password_reset_view(request):
    """Display password reset request form"""
    if request.method == 'POST':
        email = request.POST.get('email')
        try:
            user = User.objects.get(email=email)
            # Generate token and uid
            from django.contrib.auth.tokens import default_token_generator
            from django.utils.encoding import force_bytes
            from django.utils.http import urlsafe_base64_encode
            
            uid = urlsafe_base64_encode(force_bytes(user.pk))
            token = default_token_generator.make_token(user)
            
            # Send email (simplified - in production use Django's send_mail)
            reset_url = request.build_absolute_uri(
                f"/password_reset_confirm/{uid}/{token}/"
            )
            
            # For demo purposes, show the reset link directly
            # In production, send this via email
            request.session['reset_url'] = reset_url
            request.session['reset_email'] = email
            
            messages.success(request, f"Password reset link sent to {email}")
            return redirect('password_reset_done')
        except User.DoesNotExist:
            messages.error(request, "No account found with this email address")
    
    return render(request, 'password_reset.html')


def password_reset_done_view(request):
    """Display message after password reset request"""
    email = request.session.get('reset_email', '')
    return render(request, 'password_reset_done.html', {'email': email})


def password_reset_confirm_view(request, uidb64, token):
    """Handle password reset with token verification"""
    from django.contrib.auth.tokens import default_token_generator
    from django.utils.http import urlsafe_base64_decode
    
    try:
        uid = urlsafe_base64_decode(uidb64).decode()
        user = User.objects.get(pk=uid)
        
        if not default_token_generator.check_token(user, token):
            messages.error(request, "Password reset link has expired")
            return redirect('login')
        
        if request.method == 'POST':
            password1 = request.POST.get('new_password1')
            password2 = request.POST.get('new_password2')
            
            if not password1 or not password2:
                messages.error(request, "Both password fields are required")
                return render(request, 'password_reset_confirm.html')
            
            if password1 != password2:
                messages.error(request, "Passwords do not match")
                return render(request, 'password_reset_confirm.html')
            
            if len(password1) < 8:
                messages.error(request, "Password must be at least 8 characters")
                return render(request, 'password_reset_confirm.html')
            
            user.set_password(password1)
            user.save()
            messages.success(request, "Password reset successfully!")
            return redirect('password_reset_complete')
        
        return render(request, 'password_reset_confirm.html', {'uidb64': uidb64, 'token': token})
    
    except (TypeError, ValueError, OverflowError, User.DoesNotExist):
        messages.error(request, "Invalid password reset link")
        return redirect('login')


def password_reset_complete_view(request):
    """Display success message after password reset"""
    return render(request, 'password_reset_complete.html')


# ===== CART MANAGEMENT VIEWS =====

@login_required(login_url='login')
@require_POST
def add_to_cart(request):
    """Add product to cart or increase quantity"""
    product_id = request.POST.get('product_id')
    quantity = int(request.POST.get('quantity', 1))
    
    is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
    
    try:
        product = Product.objects.get(id=product_id, is_active=True)

        if product.stock <= 0:
            if is_ajax:
                return JsonResponse({
                    'success': False,
                    'message': 'This item is out of stock'
                }, status=400)
            messages.error(request, f"{product.name} is out of stock")
            return redirect('product-details', product_id=product.id)

        cart_item, created = Cart.objects.get_or_create(
            user=request.user,
            product=product,
            defaults={'quantity': quantity}
        )
        
        if not created:
            new_qty = cart_item.quantity + quantity
            if new_qty > product.stock:
                new_qty = product.stock
            cart_item.quantity = new_qty
            cart_item.save()
        
        # Get total cart count
        cart_count = Cart.objects.filter(user=request.user).count()
        
        if is_ajax:
            return JsonResponse({
                'success': True,
                'message': f'{product.name} added to cart!',
                'cart_count': cart_count,
                'product_name': product.name
            })
        else:
            messages.success(request, f"{product.name} added to cart!")
            return redirect('cart')
    except Product.DoesNotExist:
        if is_ajax:
            return JsonResponse({
                'success': False,
                'message': 'Product not found'
            }, status=404)
        else:
            messages.error(request, "Product not found")
            return redirect('shop')


@require_POST
def request_stock_notification(request, product_id):
    """Capture email to notify when a product is restocked."""
    is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
    product = get_object_or_404(Product, id=product_id)

    email = request.POST.get('email') or (request.user.email if request.user.is_authenticated else '')
    if not email:
        message = 'Please provide an email so we can notify you.'
        if is_ajax:
            return JsonResponse({'success': False, 'message': message}, status=400)
        messages.error(request, message)
        return redirect('product-details', product_id=product.id)

    notification, created = ProductStockNotification.objects.get_or_create(
        product=product,
        email=email,
        defaults={'user': request.user if request.user.is_authenticated else None}
    )

    if not created and request.user.is_authenticated and notification.user is None:
        notification.user = request.user
        notification.save(update_fields=['user'])

    message = 'We will email you as soon as this item is back in stock.' if created else 'You are already subscribed for a restock alert.'

    if is_ajax:
        return JsonResponse({'success': True, 'message': message})

    messages.success(request, message)
    next_url = request.META.get('HTTP_REFERER') or reverse('product-details', args=[product.id])
    return redirect(next_url)


@login_required(login_url='login')
def remove_from_cart(request, cart_id):
    """Remove product from cart"""
    try:
        cart_item = Cart.objects.get(id=cart_id, user=request.user)
        product_name = cart_item.product.name
        cart_item.delete()
        messages.success(request, f"{product_name} removed from cart!")
    except Cart.DoesNotExist:
        messages.error(request, "Item not found in cart")
    
    return redirect('cart')


@login_required(login_url='login')
def ajax_toggle_cart(request, product_id):
    """AJAX endpoint to toggle product in/out of cart"""
    try:
        product = Product.objects.get(id=product_id)
        if product.stock <= 0:
            return JsonResponse({
                'success': False,
                'message': 'This item is out of stock',
                'in_cart': False
            }, status=400)
        cart_item = Cart.objects.filter(user=request.user, product=product).first()
        
        if cart_item:
            # Remove from cart
            cart_item.delete()
            in_cart = False
            action = 'removed'
            message = f'{product.name} removed from cart'
        else:
            # Add to cart
            Cart.objects.create(user=request.user, product=product, quantity=1)
            in_cart = True
            action = 'added'
            message = f'{product.name} added to cart!'
        
        cart_count = Cart.objects.filter(user=request.user).count()
        
        return JsonResponse({
            'success': True,
            'in_cart': in_cart,
            'action': action,
            'message': message,
            'cart_count': cart_count
        })
    except Product.DoesNotExist:
        return JsonResponse({
            'success': False,
            'message': 'Product not found'
        }, status=404)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': str(e)
        }, status=500)


@login_required(login_url='login')
@require_POST
def update_cart_quantity(request, cart_id):
    """Update cart item quantity"""
    try:
        quantity = int(request.POST.get('quantity', 1))
        cart_item = Cart.objects.get(id=cart_id, user=request.user)
        
        if quantity <= 0:
            cart_item.delete()
            messages.success(request, "Item removed from cart")
        else:
            cart_item.quantity = quantity
            cart_item.save()
            messages.success(request, "Quantity updated")
    except (Cart.DoesNotExist, ValueError):
        messages.error(request, "Error updating quantity")
    
    return redirect('cart')


# ===== WISHLIST MANAGEMENT VIEWS =====

@login_required(login_url='login')
@require_POST
def add_to_wishlist(request):
    """Add product to wishlist"""
    product_id = request.POST.get('product_id')
    
    try:
        product = Product.objects.get(id=product_id, is_active=True)
        wishlist_item, created = Wishlist.objects.get_or_create(
            user=request.user,
            product=product
        )
        
        if created:
            messages.success(request, f"{product.name} added to wishlist!")
        else:
            messages.info(request, f"{product.name} is already in your wishlist")
        
        return redirect('wishlist')
    except Product.DoesNotExist:
        messages.error(request, "Product not found")
        return redirect('shop')


@login_required(login_url='login')
def remove_from_wishlist(request, wishlist_id):
    """Remove product from wishlist"""
    try:
        wishlist_item = Wishlist.objects.get(id=wishlist_id, user=request.user)
        product_name = wishlist_item.product.name
        wishlist_item.delete()
        messages.success(request, f"{product_name} removed from wishlist!")
    except Wishlist.DoesNotExist:
        messages.error(request, "Item not found in wishlist")
    
    return redirect('wishlist')


@login_required(login_url='login')
def move_wishlist_to_cart(request, wishlist_id):
    """Move item from wishlist to cart"""
    try:
        wishlist_item = Wishlist.objects.get(id=wishlist_id, user=request.user)
        product = wishlist_item.product
        
        # Add to cart
        cart_item, created = Cart.objects.get_or_create(
            user=request.user,
            product=product,
            defaults={'quantity': 1}
        )
        
        # Remove from wishlist
        wishlist_item.delete()
        
        messages.success(request, f"{product.name} moved to cart!")
    except Wishlist.DoesNotExist:
        messages.error(request, "Item not found in wishlist")
    
    return redirect('cart')


@login_required(login_url='login')
@require_POST
def submit_review(request, product_id):
    """Submit a product review with images (requires admin approval)"""
    try:
        product = Product.objects.get(id=product_id, is_active=True)
        
        rating = int(request.POST.get('rating', 5))
        comment = request.POST.get('comment', '').strip()
        
        # Auto-fill user details if logged in
        if request.user.is_authenticated:
            name = request.user.get_full_name() or request.user.username
            email = request.user.email
        else:
            name = request.POST.get('name', '').strip()
            email = request.POST.get('email', '').strip()
        
        # Check if user has purchased this product (verified purchase)
        is_verified_purchase = False
        order_id = request.POST.get('order_id')
        if request.user.is_authenticated and order_id:
            is_verified_purchase = Order.objects.filter(
                id=order_id,
                user=request.user,
                order_status='DELIVERED',
                items__product=product
            ).exists()
        
        # Create review (not approved by default)
        review = ProductReview.objects.create(
            product=product,
            user=request.user if request.user.is_authenticated else None,
            rating=rating,
            name=name,
            email=email,
            comment=comment,
            is_approved=False,  # Admin must approve
            is_verified_purchase=is_verified_purchase
        )
        
        # Handle image uploads (up to 5 images)
        images = request.FILES.getlist('review_images')
        for i, image in enumerate(images[:5]):  # Limit to 5 images
            ReviewImage.objects.create(review=review, image=image)
        
        messages.success(request, "Thank you for your review! It will be visible after admin approval.")
        
    except Product.DoesNotExist:
        messages.error(request, "Product not found")
    except ValueError:
        messages.error(request, "Invalid rating value")
    
    # Redirect to referrer or homepage
    return redirect(request.META.get('HTTP_REFERER', 'index'))


@login_required
@require_POST
def vote_review(request, review_id):
    """AJAX endpoint to vote on review helpfulness"""
    try:
        review = get_object_or_404(ProductReview, id=review_id, is_approved=True)
        is_helpful = request.POST.get('is_helpful') == 'true'
        
        # Check if user already voted
        existing_vote = ReviewVote.objects.filter(review=review, user=request.user).first()
        
        if existing_vote:
            # Update existing vote
            if existing_vote.is_helpful != is_helpful:
                # Decrement old count
                if existing_vote.is_helpful:
                    review.helpful_count = max(0, review.helpful_count - 1)
                else:
                    review.not_helpful_count = max(0, review.not_helpful_count - 1)
                
                # Increment new count
                if is_helpful:
                    review.helpful_count += 1
                else:
                    review.not_helpful_count += 1
                
                existing_vote.is_helpful = is_helpful
                existing_vote.save()
                review.save()
                
                return JsonResponse({
                    'success': True,
                    'message': 'Vote updated',
                    'helpful_count': review.helpful_count,
                    'not_helpful_count': review.not_helpful_count
                })
            else:
                return JsonResponse({'success': True, 'message': 'Already voted'})
        else:
            # Create new vote
            ReviewVote.objects.create(review=review, user=request.user, is_helpful=is_helpful)
            
            # Update counts
            if is_helpful:
                review.helpful_count += 1
            else:
                review.not_helpful_count += 1
            review.save()
            
            return JsonResponse({
                'success': True,
                'message': 'Vote recorded',
                'helpful_count': review.helpful_count,
                'not_helpful_count': review.not_helpful_count
            })
            
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)


@login_required
@require_POST
@login_required
def submit_question(request, product_id):
    """Submit a product question - requires admin to answer before showing on page"""
    try:
        product = get_object_or_404(Product, id=product_id, is_active=True)
        question_text = request.POST.get('question', '').strip()
        
        if not question_text:
            messages.error(request, "Please enter a question.")
            return redirect('product-details', product_id=product_id)
        
        ProductQuestion.objects.create(
            product=product,
            user=request.user,
            question=question_text,
            is_answered=False,
            is_approved=False  # Requires admin approval and answer
        )
        
        messages.success(request, "Your question has been submitted. It will appear once our team answers it!")
        
    except Exception as e:
        messages.error(request, f"Error submitting question: {str(e)}")
    
    return redirect('product-details', product_id=product_id)



from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect
from django.contrib import messages
from .models import UserProfile




from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect
from django.contrib import messages
from .models import UserProfile

@login_required(login_url='login')
def profile_view(request):

    if request.method == "POST":
        profile, created = UserProfile.objects.get_or_create(user=request.user)

        request.user.first_name = request.POST.get("first_name")
        request.user.last_name = request.POST.get("last_name")
        request.user.email = request.POST.get("email")

        profile.country_code = request.POST.get("country_code")
        profile.mobile_number = request.POST.get("mobile_number")

        if request.FILES.get("profile_image"):
            profile.profile_image = request.FILES.get("profile_image")

        request.user.save()
        profile.save()

        messages.success(request, "Profile saved successfully")

        # 🔥 THIS LINE IS REQUIRED
        return redirect('profile')

    profile, created = UserProfile.objects.get_or_create(user=request.user)
    
    # Get loyalty points and transactions
    loyalty_account, _ = LoyaltyPoints.objects.get_or_create(user=request.user)
    points_transactions = PointsTransaction.objects.filter(user=request.user).order_by('-created_at')[:10]
    
    return render(request, 'profile.html', {
        'profile': profile,
        'loyalty_account': loyalty_account,
        'points_transactions': points_transactions
    })


@login_required(login_url='login')
def api_profile_stats(request):
    """Return live profile stats and recent orders."""
    from django.http import JsonResponse
    from django.urls import reverse

    user = request.user
    orders = Order.objects.filter(user=user)

    total_orders = orders.count()
    delivered_orders = orders.filter(order_status='DELIVERED').count()
    cancelled_orders = orders.filter(order_status='CANCELLED').count()
    pending_orders = orders.filter(order_status__in=['PENDING', 'PROCESSING', 'SHIPPED']).count()

    recent_orders = orders.order_by('-created_at')[:5]
    recent_list = []
    for order in recent_orders:
        recent_list.append({
            'order_number': order.order_number,
            'status': order.order_status,
            'total_amount': float(order.total_amount),
            'created_at': order.created_at.strftime('%d %b %Y'),
            'detail_url': reverse('order_details', args=[order.order_number]),
        })

    return JsonResponse({
        'total_orders': total_orders,
        'delivered_orders': delivered_orders,
        'pending_orders': pending_orders,
        'cancelled_orders': cancelled_orders,
        'recent_orders': recent_list,
    })





# Add this to the end of views.py

@login_required(login_url='login')
def add_product(request):
    # Check if user is admin (username = 'admin')
    if request.user.username != 'FashionHub':
        messages.error(request, 'Access denied. Only FashionHub user can add products.')
        return redirect('index')
    
    if request.method == 'POST':
        try:
            # Get basic form data
            name = request.POST.get('name')
            price = request.POST.get('price')
            old_price = request.POST.get('old_price')
            stock = request.POST.get('stock')
            rating = request.POST.get('rating', 0)
            review_count = request.POST.get('review_count', 0)
            is_active = request.POST.get('is_active') == 'on'
            is_top_deal = request.POST.get('is_top_deal') == 'on'
            
            # Get new fields
            category = request.POST.get('category')
            sku = request.POST.get('sku', '')
            brand = request.POST.get('brand', '')
            description = request.POST.get('description', '')
            weight = request.POST.get('weight', '')
            color = request.POST.get('color', '')
            # Handle multiple size selections from checkboxes
            size_list = request.POST.getlist('size')
            size = ', '.join(size_list) if size_list else ''
            
            # Get images
            image = request.FILES.get('image')
            descriptionImage = request.FILES.get('descriptionImage')
            gallery_images = request.FILES.getlist('gallery_images')
            
            # Create product
            product = Product.objects.create(
                name=name,
                price=price,
                old_price=old_price if old_price else None,
                stock=stock,
                rating=rating,
                review_count=review_count,
                is_active=True,
                is_top_deal=False,
                image=image,
                descriptionImage=descriptionImage,
                category=category if category else None,
                sku=sku,
                brand=brand,
                description=description,
                weight=weight,
                color=color,
                size=size
            )
            
            # Add gallery images if provided
            for idx, gallery_image in enumerate(gallery_images, start=1):
                ProductImage.objects.create(
                    product=product,
                    image=gallery_image,
                    order=idx,
                    is_active=True
                )
            
            messages.success(request, f'Product "{product.name}" added successfully with {len(gallery_images)} gallery images!')
            return redirect('add_product')
            
        except Exception as e:
            messages.error(request, f'Error adding product: {str(e)}')
    
    # Get all products for display
    products = Product.objects.all().order_by('-id')[:30]  # Show last 30 products
    
    return render(request, 'add_product.html', {
        'products': products,
    })

# ===== AJAX WISHLIST VIEW =====

@login_required(login_url='login')
def ajax_add_to_wishlist(request, product_id):
    """AJAX endpoint to toggle product in wishlist (add/remove)"""
    if request.method == 'POST':
        try:
            product = Product.objects.get(id=product_id, is_active=True)
            
            # Check if already in wishlist
            wishlist_item = Wishlist.objects.filter(user=request.user, product=product).first()
            
            if wishlist_item:
                # Remove from wishlist
                wishlist_item.delete()
                return JsonResponse({
                    'success': True,
                    'action': 'removed',
                    'message': f'{product.name} removed from wishlist',
                    'in_wishlist': False
                })
            else:
                # Add to wishlist
                Wishlist.objects.create(user=request.user, product=product)
                return JsonResponse({
                    'success': True,
                    'action': 'added',
                    'message': f'{product.name} added to wishlist!',
                    'in_wishlist': True
                })
                
        except Product.DoesNotExist:
            return JsonResponse({
                'success': False,
                'message': 'Product not found'
            }, status=404)
    
    return JsonResponse({'success': False, 'message': 'Invalid request'}, status=400)


@login_required(login_url='login')
def check_wishlist(request, product_id):
    """Check if product is in user's wishlist"""
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        try:
            product = Product.objects.get(id=product_id)
            in_wishlist = Wishlist.objects.filter(
                user=request.user,
                product=product
            ).exists()
            
            return JsonResponse({
                'success': True,
                'in_wishlist': in_wishlist
            })
        except Product.DoesNotExist:
            return JsonResponse({
                'success': False,
                'message': 'Product not found'
            }, status=404)
    
    return JsonResponse({'success': False, 'message': 'Invalid request'}, status=400)


@login_required(login_url='login')
@require_POST
def buy_now(request, product_id):
    """Buy Now - Store product in session and redirect to checkout"""
    try:
        product = Product.objects.get(id=product_id, is_active=True)
        quantity = int(request.POST.get('quantity', 1))
        
        if quantity < 1:
            return JsonResponse({
                'success': False,
                'message': 'Invalid quantity'
            }, status=400)
        
        # Check stock
        if product.stock < quantity:
            return JsonResponse({
                'success': False,
                'message': f'Only {product.stock} items available in stock'
            }, status=400)
        
        # Store in session for checkout
        request.session['buy_now_item'] = {
            'product_id': product.id,
            'quantity': quantity,
            'price': float(product.price)
        }
        request.session.modified = True
        
        return JsonResponse({
            'success': True,
            'message': f'{product.name} - redirecting to checkout!',
            'redirect_url': '/checkout/'
        })
        
    except Product.DoesNotExist:
        return JsonResponse({
            'success': False,
            'message': 'Product not found'
        }, status=404)
    except ValueError:
        return JsonResponse({
            'success': False,
            'message': 'Invalid quantity value'
        }, status=400)


# ===== BANNER MANAGEMENT VIEWS =====

@login_required(login_url='login')
@staff_member_required(login_url='login')
def admin_banners(request):
    """Admin Banner List"""
    banners = Banner.objects.all().order_by('order')
    context = {
        'banners': banners,
    }
    return render(request, 'admin_panel/banners.html', context)

@login_required(login_url='login')
@staff_member_required(login_url='login')
def admin_add_banner(request):
    """Admin Add Banner"""
    if request.method == 'POST':
        try:
            title = request.POST.get('title')
            subtitle = request.POST.get('subtitle', '')
            badge_text = request.POST.get('badge_text', '')
            button_text = request.POST.get('button_text', '')
            button_style = request.POST.get('button_style', 'none')
            banner_type = request.POST.get('banner_type', 'LARGE')
            page_type = request.POST.get('page_type', 'HOME')
            link_url = request.POST.get('link_url', '#')
            background_color = request.POST.get('background_color', '')
            order = request.POST.get('order', 0)
            is_active = request.POST.get('is_active') == 'on'
            image = request.FILES.get('image')
            
            banner = Banner.objects.create(
                title=title,
                subtitle=subtitle,
                badge_text=badge_text,
                button_text=button_text,
                button_style=button_style,
                banner_type=banner_type,
                page_type=page_type,
                link_url=link_url,
                background_color=background_color,
                order=order,
                is_active=is_active,
                image=image
            )
            
            messages.success(request, f'Banner "{banner.title}" added successfully!')
            return redirect('admin_banners')
            
        except Exception as e:
            messages.error(request, f'Error adding banner: {str(e)}')
    
    return render(request, 'admin_panel/add_banner.html')

@login_required(login_url='login')
@staff_member_required(login_url='login')
def admin_edit_banner(request, banner_id):
    """Admin Edit Banner"""
    banner = get_object_or_404(Banner, id=banner_id)
    
    if request.method == 'POST':
        try:
            banner.title = request.POST.get('title')
            banner.subtitle = request.POST.get('subtitle', '')
            banner.badge_text = request.POST.get('badge_text', '')
            banner.button_text = request.POST.get('button_text', '')
            banner.button_style = request.POST.get('button_style', 'none')
            banner.banner_type = request.POST.get('banner_type', 'LARGE')
            banner.page_type = request.POST.get('page_type', 'HOME')
            banner.link_url = request.POST.get('link_url', '#')
            banner.background_color = request.POST.get('background_color', '')
            banner.order = request.POST.get('order', 0)
            banner.is_active = request.POST.get('is_active') == 'on'
            
            if 'image' in request.FILES:
                banner.image = request.FILES['image']
            
            banner.save()
            
            messages.success(request, f'Banner "{banner.title}" updated successfully!')
            return redirect('admin_banners')
            
        except Exception as e:
            messages.error(request, f'Error updating banner: {str(e)}')
    
    context = {
        'banner': banner,
    }
    return render(request, 'admin_panel/edit_banner.html', context)

@login_required(login_url='login')
@staff_member_required(login_url='login')
def admin_delete_banner(request, banner_id):
    """Admin Delete Banner"""
    banner = get_object_or_404(Banner, id=banner_id)
    
    if request.method == 'POST':
        banner_title = banner.title
        banner.delete()
        messages.success(request, f'Banner "{banner_title}" deleted successfully!')
    
    return redirect('admin_banners')


# ===== SLIDER MANAGEMENT VIEWS =====

@login_required(login_url='login')
@staff_member_required(login_url='login')
def admin_sliders(request):
    """Admin Slider List"""
    sliders = Slider.objects.all().order_by('order', '-id')
    context = {
        'sliders': sliders,
    }
    return render(request, 'admin_panel/sliders.html', context)

@login_required(login_url='login')
@staff_member_required(login_url='login')
def admin_add_slider(request):
    """Admin Add Slider"""
    if request.method == 'POST':
        try:
            title = request.POST.get('title')
            subtitle = request.POST.get('subtitle', '')
            top_button_text = request.POST.get('top_button_text', '')
            top_button_url = request.POST.get('top_button_url', '#')
            order = request.POST.get('order', 0)
            is_active = request.POST.get('is_active') == 'on'
            image = request.FILES.get('image')
            
            slider = Slider.objects.create(
                title=title,
                subtitle=subtitle,
                top_button_text=top_button_text,
                top_button_url=top_button_url,
                order=int(order) if order else 0,
                is_active=is_active,
                image=image
            )
            
            messages.success(request, f'Slider "{slider.title}" added successfully!')
            return redirect('admin_sliders')
            
        except Exception as e:
            messages.error(request, f'Error adding slider: {str(e)}')
    
    return render(request, 'admin_panel/add_slider.html')

@login_required(login_url='login')
@staff_member_required(login_url='login')
def admin_edit_slider(request, slider_id):
    """Admin Edit Slider"""
    slider = get_object_or_404(Slider, id=slider_id)
    
    if request.method == 'POST':
        try:
            slider.title = request.POST.get('title')
            slider.subtitle = request.POST.get('subtitle', '')
            slider.top_button_text = request.POST.get('top_button_text', '')
            slider.top_button_url = request.POST.get('top_button_url', '#')
            order = request.POST.get('order', 0)
            slider.order = int(order) if order else 0
            slider.is_active = request.POST.get('is_active') == 'on'
            
            if 'image' in request.FILES:
                slider.image = request.FILES['image']
            
            slider.save()
            
            messages.success(request, f'Slider "{slider.title}" updated successfully!')
            return redirect('admin_sliders')
            
        except Exception as e:
            messages.error(request, f'Error updating slider: {str(e)}')
    
    context = {
        'slider': slider,
    }
    return render(request, 'admin_panel/edit_slider.html', context)

@login_required(login_url='login')
@staff_member_required(login_url='login')
def admin_delete_slider(request, slider_id):
    """Admin Delete Slider"""
    slider = get_object_or_404(Slider, id=slider_id)
    
    if request.method == 'POST':
        slider_title = slider.title
        slider.delete()
        messages.success(request, f'Slider "{slider_title}" deleted successfully!')
    
    return redirect('admin_sliders')


# ===== QUESTIONS MANAGEMENT VIEWS =====

@login_required(login_url='login')
@staff_member_required(login_url='login')
def admin_questions(request):
    """Admin Questions List"""
    questions = ProductQuestion.objects.select_related('product', 'user').order_by('-created_at')
    
    # Filter by status if provided
    status = request.GET.get('status')
    if status == 'pending':
        questions = questions.filter(is_approved=False, answer__isnull=True)
    elif status == 'answered':
        questions = questions.filter(is_approved=False, answer__isnull=False)
    elif status == 'approved':
        questions = questions.filter(is_approved=True)
    
    context = {
        'questions': questions,
    }
    return render(request, 'admin_panel/questions.html', context)

@login_required(login_url='login')
@staff_member_required(login_url='login')
def admin_approve_question(request, question_id):
    """Admin Approve/Answer Question"""
    question = get_object_or_404(ProductQuestion, id=question_id)
    
    if request.method == 'POST':
        answer = request.POST.get('answer', '')
        
        if answer:
            question.answer = answer
            question.answered_by = request.user
            question.answered_at = timezone.now()
            question.is_approved = True
            question.save()
            
            messages.success(request, 'Question answered and approved!')
        else:
            messages.error(request, 'Please provide an answer')
    
    return redirect('admin_questions')

@login_required(login_url='login')
@staff_member_required(login_url='login')
def admin_delete_question(request, question_id):
    """Admin Delete Question"""
    question = get_object_or_404(ProductQuestion, id=question_id)
    
    if request.method == 'POST':
        question.delete()
        messages.success(request, 'Question deleted successfully!')
    
    return redirect('admin_questions')

@login_required(login_url='login')
@staff_member_required(login_url='login')
@require_POST
def admin_adjust_rating(request, product_id):
    """Admin Adjust Product Rating (increment/decrement)"""
    try:
        product = get_object_or_404(Product, id=product_id)
        action = request.POST.get('action')  # 'increment' or 'decrement'
        
        if action == 'increment':
            if product.rating < 5:
                product.rating = min(5, product.rating + 0.1)
                product.save()
                messages.success(request, f'Rating increased to {product.rating:.1f}')
            else:
                messages.warning(request, 'Rating is already at maximum (5.0)')
        elif action == 'decrement':
            if product.rating > 0:
                product.rating = max(0, product.rating - 0.1)
                product.save()
                messages.success(request, f'Rating decreased to {product.rating:.1f}')
            else:
                messages.warning(request, 'Rating is already at minimum (0.0)')
        
        return JsonResponse({
            'success': True,
            'rating': round(product.rating, 1)
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)

@login_required(login_url='login')
@staff_member_required(login_url='login')
@require_POST
def admin_approve_review(request, review_id):
    """Admin Approve Review"""
    from django.http import JsonResponse
    
    review = get_object_or_404(ProductReview, id=review_id)
    
    # If POST request with edited data (from modal)
    if request.method == 'POST' and request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        import json
        try:
            data = json.loads(request.body)
            edited_rating = data.get('rating')
            edited_comment = data.get('comment')
            
            # Validate and update rating
            if edited_rating and 1 <= int(edited_rating) <= 5:
                review.rating = int(edited_rating)
            
            # Update comment if provided
            if edited_comment and edited_comment.strip():
                review.comment = edited_comment.strip()
            
            review.is_approved = True
            review.save()
            
            # Update product review count and rating
            product = review.product
            product.review_count = ProductReview.objects.filter(product=product, is_approved=True).count()
            
            # Recalculate average rating
            approved_reviews = ProductReview.objects.filter(product=product, is_approved=True)
            if approved_reviews.exists():
                from django.db.models import Avg
                avg_rating = approved_reviews.aggregate(Avg('rating'))['rating__avg']
                product.rating = round(avg_rating, 1)
            
            product.save()
            
            return JsonResponse({
                'success': True,
                'message': 'Review approved successfully!'
            })
        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': str(e)
            }, status=400)
    
    # Regular approval without editing
    review.is_approved = True
    review.save()
    
    # Update product review count
    product = review.product
    product.review_count = ProductReview.objects.filter(product=product, is_approved=True).count()
    product.save()
    
    messages.success(request, 'Review approved successfully!')
    return redirect('admin_reviews')

@login_required(login_url='login')
@staff_member_required(login_url='login')
@require_POST
def admin_delete_review(request, review_id):
    """Admin Delete Review"""
    review = get_object_or_404(ProductReview, id=review_id)
    product = review.product
    review.delete()
    
    # Update product review count
    product.review_count = ProductReview.objects.filter(product=product, is_approved=True).count()
    product.save()
    
    messages.success(request, 'Review deleted successfully!')
    return redirect('admin_reviews')

@login_required(login_url='login')
@staff_member_required(login_url='login')
def admin_review_details(request, review_id):
    """Get Review Details as JSON"""
    from django.http import JsonResponse
    
    try:
        review = get_object_or_404(ProductReview, id=review_id)
        
        # Get review images
        review_images = []
        for img in review.images.all():
            review_images.append(request.build_absolute_uri(img.image.url))
        
        # Build response data
        data = {
            'id': review.id,
            'product': {
                'id': review.product.id,
                'name': review.product.name,
                'image': request.build_absolute_uri(review.product.image.url) if review.product.image else None,
            },
            'user': {
                'name': review.name,
                'email': review.email,
            },
            'rating': review.rating,
            'comment': review.comment,
            'is_verified_purchase': review.is_verified_purchase,
            'is_approved': review.is_approved,
            'created_at': review.created_at.strftime('%B %d, %Y at %I:%M %p'),
            'images': review_images,
            'helpful_count': review.helpful_count,
            'not_helpful_count': review.not_helpful_count,
        }
        
        return JsonResponse(data)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@login_required(login_url='login')
@staff_member_required(login_url='login')
def admin_add_review(request, product_id):
    """Admin Add Review to Product"""
    from django.http import JsonResponse
    
    product = get_object_or_404(Product, id=product_id)
    
    if request.method == 'POST':
        try:
            # Get admin review details
            rating = int(request.POST.get('rating', 5))
            comment = request.POST.get('comment', '').strip()
            name = request.POST.get('name', request.user.get_full_name() or request.user.username)
            email = request.POST.get('email', request.user.email)
            
            # Validate
            if not 1 <= rating <= 5:
                return JsonResponse({'success': False, 'message': 'Rating must be between 1 and 5'}, status=400)
            
            # Create review (admin reviews are auto-approved)
            review = ProductReview.objects.create(
                product=product,
                user=request.user,
                rating=rating,
                comment=comment,
                name=name,
                email=email,
                is_approved=True,
                is_verified_purchase=False  # Admin reviews are not verified purchases
            )
            
            # Update product review count and rating
            product.review_count = ProductReview.objects.filter(product=product, is_approved=True).count()
            
            # Recalculate average rating
            from django.db.models import Avg
            approved_reviews = ProductReview.objects.filter(product=product, is_approved=True)
            if approved_reviews.exists():
                avg_rating = approved_reviews.aggregate(Avg('rating'))['rating__avg']
                product.rating = round(avg_rating, 1)
            
            product.save()
            
            return JsonResponse({
                'success': True,
                'message': 'Review added successfully!',
                'review_id': review.id
            })
        
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)}, status=500)
    
    return JsonResponse({'success': False, 'message': 'Invalid request'}, status=400)


# ===== ORDER CONFIRMATION & PAYMENT VIEWS =====

@login_required(login_url='login')
def order_confirmation(request, order_id):
    """Order Confirmation Page"""
    order = get_object_or_404(Order, id=order_id, user=request.user)
    # If payment method is Razorpay and not paid, redirect to payment page
    if order.payment_method == 'RAZORPAY' and order.payment_status != 'PAID' and order.payment_status != 'PROCESSING':
        messages.warning(request, 'Please complete your payment to confirm the order.')
        return redirect('razorpay_payment', order_id=order.id)
    
    # Calculate loyalty points earned (₹1 = 33 points, 1 point = ₹0.03)
    loyalty_points_earned = int(order.total_amount * 33)
    
    context = {
        'order': order,
        'order_items': order.items.all(),
        'loyalty_points_earned': loyalty_points_earned,
    }
    return render(request, 'order_confirmation.html', context)

@login_required(login_url='login')
def razorpay_payment(request, order_id):
    """Razorpay Payment Page with OSrder Creation"""
    import razorpay
    
    order = get_object_or_404(Order, id=order_id, user=request.user)
    
    # Get Razorpay keys from settings
    razorpay_key = getattr(settings, 'RAZORPAY_KEY_ID', '')
    razorpay_secret = getattr(settings, 'RAZORPAY_KEY_SECRET', '')
    
    if not razorpay_key or not razorpay_secret:
        messages.error(request, 'Payment gateway not configured')
        return redirect('order_confirmation', order_id=order.id)
    
    try:
        # Create Razorpay client
        client = razorpay.Client(auth=(razorpay_key, razorpay_secret))

        # Amount in paise (ensure int to avoid float issues)
        amount_paise = int(order.total_amount * 100)
        if amount_paise <= 0:
            messages.error(request, 'Invalid order amount for payment')
            return redirect('order_confirmation', order_id=order.id)

        # Create Razorpay order if not already created
        if not order.razorpay_order_id:
            razorpay_order = client.order.create({
                'amount': amount_paise,
                'currency': 'INR',
                'receipt': order.order_number,
                'notes': {
                    'order_id': str(order.id),
                    'customer': order.user.username,
                    'email': order.user.email
                }
            })
            
            # Save Razorpay order ID
            order.razorpay_order_id = razorpay_order['id']
            order.save(update_fields=['razorpay_order_id'])
        
        context = {
            'order': order,
            'razorpay_key': razorpay_key,
            'razorpay_order_id': order.razorpay_order_id,
            'order_amount': amount_paise,
        }
        return render(request, 'razorpay_payment.html', context)
        
    except Exception as e:
        messages.error(request, f'Error creating payment: {str(e)}')
        return redirect('order_confirmation', order_id=order.id)


@login_required(login_url='login')
@require_POST
def razorpay_payment_success(request):
    """Handle Razorpay Payment Success"""
    try:
        import razorpay
        
        order_id = request.POST.get('order_id')
        payment_id = request.POST.get('payment_id')
        signature = request.POST.get('signature')
        
        order = get_object_or_404(Order, id=order_id, user=request.user)
        
        # Verify signature
        razorpay_secret = getattr(settings, 'RAZORPAY_KEY_SECRET', '')
        
        client = razorpay.Client(auth=(getattr(settings, 'RAZORPAY_KEY_ID', ''), razorpay_secret))
        
        params_dict = {
            'razorpay_order_id': order.razorpay_order_id,
            'razorpay_payment_id': payment_id,
            'razorpay_signature': signature
        }
        
        try:
            client.utility.verify_payment_signature(params_dict)
            
            # Payment verified
            order.payment_status = 'PAID'
            order.order_status = 'PROCESSING'
            order.razorpay_payment_id = payment_id
            order.razorpay_signature = signature
            order.save(update_fields=['payment_status', 'order_status', 'razorpay_payment_id', 'razorpay_signature'])
            
            # Auto-process approval (fraud detection & auto-approve)
            order.auto_process_approval()
            
            # Clear cart
            Cart.objects.filter(user=request.user).delete()
            
            # Send order confirmation email
            send_order_confirmation_email(order)
            
            if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                return JsonResponse({'success': True, 'redirect': reverse('order_confirmation', args=[order.id])})
            
            # Check if order needs approval
            if order.approval_status == 'PENDING_APPROVAL':
                messages.warning(request, 'Payment successful! Your order is pending approval due to security checks.')
            else:
                messages.success(request, 'Payment successful! Your order is being processed.')
            
            return redirect('order_confirmation', order_id=order.id)
            
        except Exception as e:
            order.payment_status = 'FAILED'
            order.save(update_fields=['payment_status'])
            if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                return JsonResponse({'success': False, 'message': 'Payment verification failed'}, status=400)
            messages.error(request, 'Payment verification failed')
            return redirect('checkout')
            
    except Exception as e:
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({'success': False, 'message': f'Error processing payment: {str(e)}'}, status=400)
        messages.error(request, f'Error processing payment: {str(e)}')
        return redirect('checkout')


@login_required(login_url='login')
def razorpay_payment_cancel(request, order_id):
    """Handle Razorpay Payment Cancellation"""
    order = get_object_or_404(Order, id=order_id, user=request.user)
    
    # Update order status
    from django.utils import timezone
    order.payment_status = 'FAILED'
    order.order_status = 'CANCELLED'
    order.save()
    
    # Send cancellation email
    try:
        from django.template.loader import render_to_string
        from django.core.mail import EmailMultiAlternatives
        
        subject = f'Order Cancelled - {order.order_number}'
        
        # Get site URL
        site_url = request.build_absolute_uri('/').rstrip('/')
        
        # Render HTML email
        html_content = render_to_string('emails/order_cancelled.html', {
            'order': order,
            'cancelled_at': timezone.now(),
            'site_url': site_url,
        })
        
        # Plain text version
        text_content = f'''
Dear {order.user.get_full_name() or order.user.username},

Your order {order.order_number} has been cancelled as per your request.

Order Details:
- Order Number: {order.order_number}
- Total Amount: ₹{order.total_amount}
- Status: Cancelled

Thank you for your understanding.

Best regards,
FashioHub Team
        '''
        
        # Send email
        email = EmailMultiAlternatives(subject, text_content, settings.EMAIL_HOST_USER, [order.user.email])
        email.attach_alternative(html_content, "text/html")
        email.send()
        
    except Exception as e:
        print(f'Email sending failed: {e}')
    
    messages.warning(request, 'Payment was cancelled. You can try again from your orders.')
    return redirect('checkout')


from django.views.decorators.csrf import csrf_exempt


@csrf_exempt
@require_POST
def razorpay_webhook(request):
    """Handle Razorpay Webhook Events (no auth, signature verified)"""
    import razorpay
    import json
    from django.views.decorators.csrf import csrf_exempt
    
    webhook_secret = getattr(settings, 'RAZORPAY_WEBHOOK_SECRET', '')
    webhook_signature = request.headers.get('X-Razorpay-Signature', '')
    webhook_body = request.body
    
    if not webhook_secret:
        return JsonResponse({'status': 'webhook not configured'}, status=400)
    
    # Verify webhook signature
    try:
        # Verify signature
        import hmac
        import hashlib
        
        expected_signature = hmac.new(
            webhook_secret.encode('utf-8'),
            webhook_body,
            hashlib.sha256
        ).hexdigest()
        
        if expected_signature != webhook_signature:
            return JsonResponse({'status': 'invalid signature'}, status=400)
            
    except Exception as e:
        return JsonResponse({'status': 'verification failed'}, status=400)
    
    # Process webhook data
    try:
        data = json.loads(webhook_body)
        event = data.get('event')
        
        if event == 'payment.captured':
            # Payment successful
            payment = data['payload']['payment']['entity']
            order_id = payment['notes'].get('order_id')
            
            if order_id:
                order = Order.objects.get(id=order_id)
                order.payment_status = 'PAID'
                order.order_status = 'PROCESSING'
                order.razorpay_payment_id = payment['id']
                order.save()
                
                # Clear user's cart
                Cart.objects.filter(user=order.user).delete()
                
        elif event == 'payment.failed':
            # Payment failed
            payment = data['payload']['payment']['entity']
            order_id = payment['notes'].get('order_id')
            
            if order_id:
                order = Order.objects.get(id=order_id)
                order.payment_status = 'FAILED'
                order.save()
        
        return JsonResponse({'status': 'ok'})
        
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)


@login_required(login_url='login')
@user_passes_test(lambda u: u.is_staff)
def razorpay_refund(request, order_id):
    """Process Razorpay Refund for an Order"""
    import razorpay
    
    order = get_object_or_404(Order, id=order_id)
    
    # Check if order can be refunded
    if order.payment_status != 'PAID':
        messages.error(request, 'Only paid orders can be refunded')
        return redirect('admin_order_details', order_id=order.id)
    
    if not order.razorpay_payment_id:
        messages.error(request, 'No payment ID found for this order')
        return redirect('admin_order_details', order_id=order.id)
    
    # Get Razorpay credentials
    razorpay_key = getattr(settings, 'RAZORPAY_KEY_ID', '')
    razorpay_secret = getattr(settings, 'RAZORPAY_KEY_SECRET', '')
    
    if not razorpay_key or not razorpay_secret:
        messages.error(request, 'Payment gateway not configured')
        return redirect('admin_order_details', order_id=order.id)
    
    try:
        # Create Razorpay client
        client = razorpay.Client(auth=(razorpay_key, razorpay_secret))
        
        # Get refund amount (can be partial or full)
        refund_amount = request.POST.get('refund_amount')
        if refund_amount:
            amount_in_paise = int(float(refund_amount) * 100)
        else:
            # Full refund
            amount_in_paise = int(order.total_amount * 100)
        
        # Create refund
        refund = client.payment.refund(
            order.razorpay_payment_id,
            {
                'amount': amount_in_paise,
                'speed': 'normal',
                'notes': {
                    'reason': request.POST.get('refund_reason', 'Customer requested refund'),
                    'order_number': order.order_number,
                    'refunded_by': request.user.username
                },
                'receipt': f'REFUND_{order.order_number}'
            }
        )
        
        # Update order status
        order.payment_status = 'REFUNDED'
        order.order_status = 'CANCELLED'
        order.admin_notes = f"Refund ID: {refund['id']}\nRefund Amount: ₹{refund_amount or order.total_amount}\n{order.admin_notes}"
        order.save()
        
        messages.success(request, f'Refund initiated successfully! Refund ID: {refund["id"]}')
        
    except Exception as e:
        messages.error(request, f'Refund failed: {str(e)}')
    
    return redirect('admin_order_details', order_id=order.id)


@login_required(login_url='login')
def resell_order(request, order_id):
    """Create a Resell Order from existing order"""
    try:
        original_order = get_object_or_404(Order, id=order_id, user=request.user)
        
        # Create new order with is_resell=True
        new_order = Order.objects.create(
            user=request.user,
            subtotal=original_order.subtotal,
            tax=original_order.tax,
            shipping_cost=original_order.shipping_cost,
            total_amount=original_order.total_amount,
            shipping_address=original_order.shipping_address,
            billing_address=original_order.billing_address,
            payment_method=original_order.payment_method,
            is_resell=True
        )
        
        # Copy order items
        for item in original_order.items.all():
            OrderItem.objects.create(
                order=new_order,
                product=item.product,
                product_name=item.product_name,
                product_price=item.product_price,
                product_image=item.product_image,
                quantity=item.quantity,
                size=item.size,
                color=item.color
            )
        
        messages.success(request, f'Resell order created! Order #: {new_order.order_number}')
        return redirect('checkout')
        
    except Order.DoesNotExist:
        messages.error(request, 'Order not found')
        return redirect('profile')


def order_list(request):
    """Display list of user's orders"""
    if not request.user.is_authenticated:
        return redirect('login')
    
    orders = Order.objects.filter(user=request.user).order_by('-created_at')
    return render(request, 'order_list.html', {'orders': orders})


def track_order_page(request):
    """Order tracking search page"""
    if request.method == 'POST':
        order_number = request.POST.get('order_number', '').strip()
        if order_number:
            return redirect('order_tracking', order_number=order_number)
        else:
            messages.error(request, 'Please enter an order number')
    
    return render(request, 'track_order.html')


def order_details(request, order_number):
    """Display detailed view of a specific order"""
    if not request.user.is_authenticated:
        return redirect('login')
    
    try:
        order = Order.objects.get(order_number=order_number, user=request.user)
        order_items = OrderItem.objects.filter(order=order)
        return render(request, 'order_details.html', {
            'order': order,
            'order_items': order_items
        })
    except Order.DoesNotExist:
        messages.error(request, 'Order not found')
        return redirect('order_list')


def order_tracking(request, order_number):
    """Display beautiful timeline tracking for a specific order"""
    if not request.user.is_authenticated:
        return redirect('login')
    
    try:
        order = Order.objects.get(order_number=order_number, user=request.user)
        order_items = OrderItem.objects.filter(order=order)
        return render(request, 'order_tracking.html', {
            'order': order,
            'order_items': order_items
        })
    except Order.DoesNotExist:
        messages.error(request, 'Order not found')
        return redirect('order_list')


@login_required(login_url='login')
@require_POST
def customer_cancel_order(request, order_id):
    """Allow customer to cancel their own order (before shipping)"""
    order = get_object_or_404(Order, id=order_id, user=request.user)
    
    # Only allow cancellation if order is PENDING or PROCESSING
    if order.order_status not in ['PENDING', 'PROCESSING', 'PENDING_APPROVAL']:
        messages.error(request, 'This order cannot be cancelled as it has already been shipped.')
        return redirect(request.META.get('HTTP_REFERER', 'index'))
    
    # Update order status
    from django.utils import timezone
    old_status = order.order_status
    order.order_status = 'CANCELLED'
    order.approval_notes = f'Customer cancelled order on {timezone.now().strftime("%d %b, %Y at %I:%M %p")}'
    order.save()
    
    # Send cancellation email
    try:
        from django.template.loader import render_to_string
        from django.core.mail import EmailMultiAlternatives
        
        subject = f'Order Cancelled - {order.order_number}'
        
        # Get site URL
        site_url = request.build_absolute_uri('/').rstrip('/')
        
        # Render HTML email
        html_content = render_to_string('emails/order_cancelled.html', {
            'order': order,
            'cancelled_at': timezone.now(),
            'site_url': site_url,
        })
        
        # Plain text version
        text_content = f'''
Dear {order.user.get_full_name() or order.user.username},

Your order {order.order_number} has been cancelled as per your request.

Order Details:
- Order Number: {order.order_number}
- Total Amount: ₹{order.total_amount}
- Status: Cancelled
- Cancelled On: {timezone.now().strftime("%d %b, %Y")}

{"A refund will be initiated within 5-7 business days." if order.payment_status == "Paid" else "No payment has been deducted from your account."}

Thank you for your understanding.

Best regards,
FashioHub Team
        '''
        
        # Send email
        email = EmailMultiAlternatives(subject, text_content, settings.EMAIL_HOST_USER, [order.user.email])
        email.attach_alternative(html_content, "text/html")
        email.send()
        
    except Exception as e:
        print(f'Email sending failed: {e}')
    
    messages.success(request, f'Order {order.order_number} has been cancelled successfully.')
    return redirect(request.META.get('HTTP_REFERER', 'index'))


@login_required(login_url='login')
@never_cache
def download_invoice(request, order_number):
    """Generate and download PDF invoice for an order"""
    import os
    import traceback
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak, Image
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT, TA_JUSTIFY
    from io import BytesIO
    
    # Register HEIF/AVIF support for Pillow
    try:
        import pillow_heif
        pillow_heif.register_heif_opener()
    except ImportError:
        pass  # pillow-heif not installed, will fall back to showing '-'
    
    try:
        # Get order
        order_filter = {'order_number': order_number}
        if not request.user.is_staff:
            order_filter['user'] = request.user

        try:
            order = Order.objects.get(**order_filter)
            order_items = OrderItem.objects.filter(order=order)
        except Order.DoesNotExist:
            # Don't use messages here; just return error response
            if request.user.is_staff:
                return redirect('admin_invoice_inventory')
            return HttpResponse('Order not found.', status=404)
        
        # Create PDF buffer
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, 
                                rightMargin=36, leftMargin=36,
                                topMargin=36, bottomMargin=30)
        
        elements = []
        styles = getSampleStyleSheet()
        
        # Define custom styles
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#1e293b'),
            spaceAfter=6,
            fontName='Helvetica-Bold'
        )
        
        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading2'],
            fontSize=11,
            textColor=colors.HexColor('#1e293b'),
            spaceAfter=8,
            fontName='Helvetica-Bold',
            borderPadding=6,
            backgroundColor=colors.HexColor('#f8fafc'),
            borderColor=colors.HexColor('#e2e8f0'),
            borderWidth=0.5,
            leftIndent=6
        )
        
        label_style = ParagraphStyle(
            'Label',
            parent=styles['Normal'],
            fontSize=9,
            textColor=colors.HexColor('#64748b'),
            fontName='Helvetica-Bold'
        )
        
        value_style = ParagraphStyle(
            'Value',
            parent=styles['Normal'],
            fontSize=10,
            textColor=colors.HexColor('#1e293b'),
            fontName='Helvetica'
        )
        
        # ===== HEADER SECTION =====
        header_data = [
            [
                Paragraph('<b>VibeMall</b>', styles['Heading2']),
                Paragraph(f'<b style="font-size: 18">INVOICE</b>', styles['Heading2'])
            ]
        ]
        header_table = Table(header_data, colWidths=[4*inch, 2*inch])
        header_table.setStyle(TableStyle([
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('ALIGN', (0, 0), (0, 0), 'LEFT'),
            ('ALIGN', (1, 0), (1, 0), 'RIGHT'),
        ]))
        elements.append(header_table)
        elements.append(Spacer(1, 12))
        
        # ===== INVOICE INFO ROW =====
        # Ensure a font that supports the rupee symbol is registered.
        # Try several common system font locations and register the first available TTF.
        try:
            from reportlab.pdfbase import pdfmetrics
            from reportlab.pdfbase.ttfonts import TTFont
            candidate_paths = [
                # Common Linux paths
                '/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf',
                '/usr/share/fonts/truetype/noto/NotoSans-Regular.ttf',
                '/usr/share/fonts/truetype/freefont/FreeSans.ttf',
                # Common Windows paths
                'C:/Windows/Fonts/DejaVuSans.ttf',
                'C:/Windows/Fonts/NotoSans-Regular.ttf',
                'C:/Windows/Fonts/seguiemj.ttf',
                'C:/Windows/Fonts/SegoeUIEmoji.ttf',
                'C:/Windows/Fonts/arialuni.ttf',
                'C:/Windows/Fonts/arial.ttf',
            ]

            base_font_name = None
            bold_font_name = None
            for path in candidate_paths:
                if os.path.exists(path):
                    try:
                        pdfmetrics.registerFont(TTFont('Invoice', path))
                        base_font_name = 'Invoice'
                        # try to find a bold variant nearby
                        bold_candidates = [
                            path.replace('.ttf', ' Bold.ttf'),
                            path.replace('.ttf', '-Bold.ttf'),
                            path.replace('.ttf', 'Bd.ttf'),
                            path.replace('.ttf', 'Bold.ttf'),
                        ]
                        for bpath in bold_candidates:
                            if os.path.exists(bpath):
                                try:
                                    pdfmetrics.registerFont(TTFont('Invoice-Bold', bpath))
                                    bold_font_name = 'Invoice-Bold'
                                    break
                                except Exception:
                                    continue
                        if not bold_font_name:
                            # fallback to same font for bold if no bold file found
                            bold_font_name = base_font_name
                        break
                    except Exception:
                        continue

            if not base_font_name:
                base_font_name = 'Helvetica'
                bold_font_name = 'Helvetica-Bold'
        except Exception:
            base_font_name = 'Helvetica'
            bold_font_name = 'Helvetica-Bold'

        # apply chosen fonts to styles
        value_style.fontName = base_font_name
        title_style.fontName = bold_font_name
        heading_style.fontName = bold_font_name
        label_style.fontName = bold_font_name

        invoice_info = f"""
        <b>Invoice #:</b> {order.order_number}<br/>
        <b>Date:</b> {order.created_at.strftime('%d %B %Y')}<br/>
        <b>Time:</b> {order.created_at.strftime('%H:%M %p')}
        """

        # Keep Order Status and Payment Method but remove Payment Status per request
        order_info = f"""
        <b>Order Status:</b> {order.get_order_status_display()}<br/>
        <b>Payment Method:</b> {order.payment_method.upper()}
        """

        info_data = [
            [Paragraph(invoice_info, value_style), Paragraph(order_info, value_style)]
        ]
        info_table = Table(info_data, colWidths=[3*inch, 3*inch])
        info_table.setStyle(TableStyle([
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('LEFTPADDING', (0, 0), (-1, -1), 0),
            ('RIGHTPADDING', (0, 0), (-1, -1), 0),
        ]))
        elements.append(info_table)
        elements.append(Spacer(1, 14))
        
        # ===== BILLING & SHIPPING SECTION =====
        # Log invoice generation start
        try:
            logger = logging.getLogger(__name__)
            logger.info("Generating invoice PDF for order %s user=%s", order_number, request.user.username if hasattr(request.user, 'username') else str(request.user))
        except Exception:
            pass

        billing_data = [
            [
                Paragraph('<b style="font-size: 11; color: #1e293b">BILL TO</b>', styles['Normal']),
                Paragraph('<b style="font-size: 11; color: #1e293b">SHIP TO</b>', styles['Normal'])
            ],
            [
                Paragraph(
                    f'<b>{order.user.get_full_name() or order.user.username}</b><br/>'
                    f'{order.user.email}<br/>'
                    f'{order.user.userprofile.mobile_number if hasattr(order.user, "userprofile") and order.user.userprofile.mobile_number else "N/A"}',
                    value_style
                ),
                Paragraph((order.shipping_address or '').replace('\n', '<br/>'), value_style)
            ]
        ]
        billing_table = Table(billing_data, colWidths=[3*inch, 3*inch])
        billing_table.setStyle(TableStyle([
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#f8fafc')),
            ('LINEBELOW', (0, 0), (-1, 0), 1, colors.HexColor('#e2e8f0')),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ]))
        elements.append(billing_table)
        elements.append(Spacer(1, 14))
        
        # ===== ITEMS TABLE ===== (with product image and name combined)
        items_data = [[ 'Product', 'Qty', 'Unit Price', 'Total' ]]

        for item in order_items:
            item_total = Decimal(str(item.product_price)) * Decimal(str(item.quantity))

            # product image and name cell (merged)
            product_cell = None
            img_cell = None
            
            if item.product_image and str(item.product_image).strip():
                img_path_str = str(item.product_image).strip()
                # ignore external URLs
                if not img_path_str.startswith('http'):
                    # use MEDIA_ROOT if configured, otherwise fallback to BASE_DIR/media
                    media_root = getattr(settings, 'MEDIA_ROOT', None) or os.path.join(settings.BASE_DIR, 'media')
                    # Normalize common prefixes: '/media/', 'media/', or leading '/'
                    if img_path_str.startswith('/media/'):
                        img_path_str = img_path_str[len('/media/'):]
                    elif img_path_str.startswith('media/'):
                        img_path_str = img_path_str[len('media/'):]
                    elif img_path_str.startswith('/'):
                        img_path_str = img_path_str[1:]
                    img_path = os.path.join(media_root, img_path_str)
                    if img_path and os.path.exists(img_path) and os.path.getsize(img_path) > 0:
                        # Try to load image with PIL (with AVIF support via pillow-heif)
                        try:
                            from PIL import Image as PILImage
                            # Open image to verify it's readable
                            with open(img_path, 'rb') as f:
                                test_img = PILImage.open(f)
                                test_img.load()  # Force load to catch errors
                            # Image is readable, use it
                            img_cell = Image(img_path, width=0.5*inch, height=0.5*inch)
                        except Exception:
                            # Image format not supported, show dash
                            img_cell = None
            
            # Create product cell with image (if available) and name
            if img_cell:
                # Use a nested table: image on left, name on right
                from reportlab.platypus import Table as NestedTable, TableStyle as NestedTableStyle
                product_inner_data = [[img_cell, Paragraph(item.product_name[:35], value_style)]]
                product_nested = NestedTable(product_inner_data, colWidths=[0.55*inch, 2.45*inch])
                product_nested.setStyle(NestedTableStyle([
                    ('ALIGN', (0, 0), (0, 0), 'CENTER'),
                    ('ALIGN', (1, 0), (1, 0), 'LEFT'),
                    ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                    ('LEFTPADDING', (0, 0), (-1, -1), 0),
                    ('RIGHTPADDING', (0, 0), (-1, -1), 0),
                    ('TOPPADDING', (0, 0), (-1, -1), 0),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 0),
                ]))
                product_cell = product_nested
            else:
                # No image, just show name
                product_cell = Paragraph(item.product_name[:40], value_style)

            items_data.append([
                product_cell,
                str(item.quantity),
                f"Rs.{float(item.product_price):.2f}",
                f"Rs.{float(item_total):.2f}",
            ])

        items_table = Table(items_data, colWidths=[3.0*inch, 0.6*inch, 1.0*inch, 1.0*inch])
        items_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#f8fafc')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#1e293b')),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), bold_font_name),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
            ('TOPPADDING', (0, 0), (-1, 0), 8),
            ('LINEBELOW', (0, 0), (-1, 0), 1, colors.HexColor('#e2e8f0')),
            ('ALIGN', (0, 1), (-1, -1), 'CENTER'),
            ('ALIGN', (0, 1), (0, -1), 'LEFT'),
            ('ALIGN', (1, 1), (1, -1), 'CENTER'),
            ('ALIGN', (2, 1), (-1, -1), 'RIGHT'),
            ('VALIGN', (0, 1), (-1, -1), 'MIDDLE'),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('TOPPADDING', (0, 1), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 8),
            ('LINEBELOW', (0, 1), (-1, -1), 0.5, colors.HexColor('#e2e8f0')),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e2e8f0')),
        ]))
        elements.append(items_table)
        elements.append(Spacer(1, 12))
        
        # ===== TOTALS SECTION =====
        tax_amount = float(order.tax)
        subtotal = float(order.subtotal)
        shipping = float(order.shipping_cost)
        total = float(order.total_amount)
        
        totals_data = [
            ['Subtotal', '', f'Rs.{subtotal:.2f}'],
            ['Tax', '', f'Rs.{tax_amount:.2f}'],
            ['Shipping', '', f'Rs.{shipping:.2f}'],
            ['TOTAL', '', f'Rs.{total:.2f}'],
        ]
        
        totals_table = Table(totals_data, colWidths=[2.5*inch, 1.5*inch, 2.0*inch])
        totals_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (0, -1), 'LEFT'),
            ('ALIGN', (2, 0), (2, -1), 'RIGHT'),
            ('FONTSIZE', (0, 0), (-1, 2), 9),
            ('FONTSIZE', (0, 3), (-1, 3), 11),
            ('FONTNAME', (0, 0), (-1, 2), base_font_name),
            ('FONTNAME', (0, 3), (-1, 3), bold_font_name),
            ('TOPPADDING', (0, 0), (-1, 2), 8),
            ('BOTTOMPADDING', (0, 0), (-1, 2), 8),
            ('TOPPADDING', (0, 3), (-1, 3), 12),
            ('BOTTOMPADDING', (0, 3), (-1, 3), 12),
            ('LINEABOVE', (0, 3), (-1, 3), 2, colors.HexColor('#1e293b')),
            ('BACKGROUND', (0, 3), (-1, 3), colors.HexColor('#f8fafc')),
            ('TEXTCOLOR', (0, 0), (-1, 2), colors.HexColor('#475569')),
            ('TEXTCOLOR', (0, 3), (-1, 3), colors.HexColor('#1e293b')),
        ]))
        elements.append(totals_table)
        elements.append(Spacer(1, 16))
        
        # ===== NOTES SECTION =====
        elements.append(Paragraph('<b style="font-size: 10; color: #1e293b">ORDER NOTES</b>', styles['Normal']))
        elements.append(Spacer(1, 6))
        
        notes_text = f"""
        <font size="9" color="#475569">
        <b>Customer Notes:</b> {order.customer_notes or 'No special notes'}<br/><br/>
        Thank you for your purchase! Your order is being processed and will be shipped soon.
        </font>
        """
        elements.append(Paragraph(notes_text, value_style))
        elements.append(Spacer(1, 12))
        
        # ===== FOOTER =====
        footer_text = f"""
        <font size="8" color="#64748b">
        <b>VibeMall © 2026</b> | Invoice #{order.order_number} | Generated on {order.created_at.strftime('%d %b %Y at %H:%M')}
        </font>
        """
        elements.append(Paragraph(footer_text, styles['Normal']))
        
        # Build PDF
        doc.build(elements)
        pdf_data = buffer.getvalue()
        buffer.close()
        
        # Return PDF with proper headers
        response = HttpResponse(pdf_data, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="Invoice_{order.order_number}.pdf"'
        response['Content-Length'] = len(pdf_data)
        response['Cache-Control'] = 'no-cache, no-store, must-revalidate'
        response['Pragma'] = 'no-cache'
        response['Expires'] = '0'
        return response
        
    except Exception as e:
        logger = logging.getLogger(__name__)
        logger.exception("Error generating invoice for %s: %s", order_number, str(e))
        # Return error response without trying to use messages
        return HttpResponse('Error generating invoice. Please contact support.', status=500)
